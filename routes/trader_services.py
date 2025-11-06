# routes/trader_services.py
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, current_app, jsonify
from flask_login import login_required, current_user
from utils.decorators import role_required, permission_required
from models_reports import ReportState
from models import db
import pandas as pd
import json, io, re
from decimal import Decimal, InvalidOperation

trader_services_bp = Blueprint("trader_services_bp", __name__)

SECTIONS = {
    "frequent": "المترددين",
    "primary": "الماكينات الأساسية للعملاء وماكينات الفرع",
}

# قائمة أنواع الأعطال المتاحة (نفس القائمة من machine_reports.py)
FAULT_TYPES = ["ريدر","سوفت","طباعه","شحن","سوكت","شبكه","شاشه","بيت شريحه","F2","KEYS","POWER"]

# الأعمدة التي يجب استبعادها من قائمة البحث في شاشة المترددين
# نضيف أيضًا "ملاحظات1" كمرادف يُستخدم أحيانًا بدل "ملاحظات"
EXCLUDED_SEARCH_COLUMNS = set(FAULT_TYPES + ["الحوالة المطلوبة", "ملاحظات", "ملاحظات1"]) 

# تطبيع اسم العمود للمقارنة المتسقة (إزالة المسافات وتحويل لحروف صغيرة)
def _normalize_name(s: str) -> str:
    try:
        return re.sub(r"\s+", "", str(s).strip().lower())
    except Exception:
        return str(s).strip().lower()

# مجموعة مطبّعة للأسماء المستبعدة لمطابقة اختلافات طفيفة في الكتابة
EXCLUDED_SEARCH_COLUMNS_NORM = { _normalize_name(x) for x in EXCLUDED_SEARCH_COLUMNS }

# الترتيب الافتراضي المتفق عليه لعرض المترددين
# ملاحظة: ندرج بعض البدائل لضمان ظهور الأعمدة الشائعة إن كانت بأسماء مختلفة.
# الأولوية تكون للاسم المطلوب ثم البدائل.
DEFAULT_FREQUENT_ORDER = [
    'النوع', 'التوع',
    'الادارة',
    'مكتب', 'المكتب',
    'اسم العميل', 'رقم العميل',
    'مسلسل', 'التاريخ',
    'خدمات',
    'صيانه', 'صيانة',
    'الاذن', 'رقم الإذن', 'رقم الاذن',
    # أنواع الأعطال بالترتيب (بنفس الترتيب المطلوب)
    *FAULT_TYPES,
    'الحوالة المطلوبة',
    'ملاحظات', 'ملاحظات1',
    '_الفترة'
]

_EMPTY_TOKENS = {"nan", "none", "null", "na", "n/a", "nat", "-", "—"}
_DATE_RE_1 = re.compile(r"^(\d{4}[-/]\d{1,2}[-/]\d{1,2})[ T]\d{1,2}:\d{2}(?::\d{2})?$")
_DATE_RE_2 = re.compile(r"^(\d{1,2}[-/]\d{1,2}[-/]\d{4})[ T]\d{1,2}:\d{2}(?::\d{2})?$")

def _strip_time_from_date(txt: str) -> str:
    m = _DATE_RE_1.match(txt)
    if m: return m.group(1)
    m = _DATE_RE_2.match(txt)
    if m: return m.group(1)
    return txt

def _textify(v) -> str:
    if v is None: return ""
    t = str(v).strip()
    if t == "" or t.lower() in _EMPTY_TOKENS: return ""
    t = t.replace(",", "")
    # عدم إزالة وقت التاريخ لضمان ظهور وقت الزيارة في الشاشات
    if re.fullmatch(r"\d+", t): return t
    if re.fullmatch(r"\d+\.\d+", t): return t.split(".", 1)[0]
    if re.fullmatch(r"[0-9]+(\.[0-9]+)?[eE][+\-]?[0-9]+", t):
        try: return str(Decimal(t).to_integral_value(rounding="ROUND_DOWN"))
        except InvalidOperation: return ""
    return t

def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip().replace("\n"," ").replace("\r"," ") for c in df.columns]
    return df

def _coerce_all_text_no_decimals(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty: return pd.DataFrame()
    out = df.copy()
    for c in out.columns: out[c] = out[c].map(_textify)
    return out

def _read_excel(file_storage) -> pd.DataFrame:
    df = pd.read_excel(file_storage, engine="openpyxl", dtype=str)
    df = _normalize_columns(df)
    df = _coerce_all_text_no_decimals(df)
    return df

def _df_to_json(df: pd.DataFrame) -> str:
    return _coerce_all_text_no_decimals(df).to_json(orient="records", force_ascii=False)

def _json_to_df(js: str) -> pd.DataFrame:
    return _coerce_all_text_no_decimals(pd.DataFrame(json.loads(js))) if js else pd.DataFrame()

def _load_state(key: str):
    return ReportState.query.filter_by(category=key).first()

def _save_state(key: str, df: pd.DataFrame=None, mapping: dict=None):
    row = _load_state(key)
    if not row:
        row = ReportState(category=key); db.session.add(row)
    if df is not None: row.data_json = _df_to_json(df)
    if mapping is not None: row.mapping_json = json.dumps(mapping, ensure_ascii=False)
    db.session.commit()

def _apply_mapping(df: pd.DataFrame, mapping: dict) -> pd.DataFrame:
    if df is None or df.empty or not mapping:
        return df
    rename = mapping.get("rename") or {}
    order  = [c for c in (mapping.get("order") or []) if c]
    out = df.copy()
    if rename:
        out = out.rename(columns=rename)
    if order:
        front = [c for c in order if c in out.columns]
        out = out[front] if front else out
    return _coerce_all_text_no_decimals(out)

def _enforce_frequent_default_order(df: pd.DataFrame) -> pd.DataFrame:
    """فرض ترتيب الأعمدة المتفق عليه لشاشة المترددين بغض النظر عن ترتيب الملف.
    - يبقي الأعمدة غير الموجودة كما هي في نهاية الجدول.
    - لا يضيف أعمدة جديدة إذا لم تكن موجودة.
    """
    if df is None or df.empty:
        return df
    cols = list(df.columns)
    front = [c for c in DEFAULT_FREQUENT_ORDER if c in cols]
    tail = [c for c in cols if c not in front]
    try:
        return df[front + tail]
    except Exception:
        return df

def _filter_dataframe(df: pd.DataFrame, q: str, search_in: str | None) -> pd.DataFrame:
    """تصفية سريعة: في حالة البحث في كل الأعمدة، نبني نصًا مجمّعًا لكل صف مرة واحدة.
    هذا يقلّل الحسابات مقارنةً بتطبيق البحث على كل عمود على حدة."""
    if not q:
        return df
    ql = (q or '').strip().lower()
    if not ql:
        return df

    # بحث في عمود محدد
    if search_in and search_in != "all" and search_in in df.columns:
        mask = df[search_in].astype(str).str.lower().str.contains(ql, na=False)
        return df[mask]

    # بحث شامل: تجميع صف واحد لنص واحد ثم contains
    try:
        cols = list(df.columns)
        # نستخدم fillna لتجنب 'nan' في التجميع
        blob = df[cols].fillna("").astype(str).agg("|".join, axis=1).str.lower()
        mask = blob.str.contains(ql, na=False)
        return df[mask]
    except Exception:
        # رجوع للخوارزمية السابقة عند حدوث أي خطأ
        mask = df.apply(lambda col: col.astype(str).str.lower().str.contains(ql, na=False))
        return df[mask.any(axis=1)]

def _drop_empty_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty: return df
    def _is_empty_series(s: pd.Series) -> bool:
        vals = s.fillna("").astype(str).str.strip().str.lower()
        return ((vals == "") | (vals.isin(_EMPTY_TOKENS))).all()
    keep = [c for c in df.columns if not _is_empty_series(df[c])]
    return df[keep]

def _paginate(df: pd.DataFrame, page: int, page_size: int):
    n = len(df); start = max(0, (page-1)*page_size); end = start + page_size
    return df.iloc[start:end], n

def _excel_response(out_df: pd.DataFrame, filename: str):
    output = io.BytesIO()
    try:
        import xlsxwriter
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            sheet = "data"; out_df.to_excel(writer, index=False, sheet_name=sheet)
            ws = writer.sheets[sheet]; book = writer.book
            header_fmt = book.add_format({"bold": True, "bg_color": "#E2E8F0", "align":"center", "valign":"vcenter", "border":1, "num_format":"@"})
            cell_fmt   = book.add_format({"align":"center", "valign":"vcenter", "border":1, "num_format":"@"})
            for col_idx, col_name in enumerate(out_df.columns): ws.write(0, col_idx, col_name, header_fmt)
            n_rows, n_cols = len(out_df.index), len(out_df.columns)
            if n_rows>0 and n_cols>0:
                from xlsxwriter.utility import xl_rowcol_to_cell
                start = xl_rowcol_to_cell(1,0); end = xl_rowcol_to_cell(n_rows, n_cols-1)
                ws.conditional_format(f"{start}:{end}", {"type":"no_blanks", "format":cell_fmt})
            for i, c in enumerate(out_df.columns):
                series = out_df[c].astype(str)
                w = min(max([len(str(c))] + [len(s) for s in series.tolist()]) + 2, 60)
                ws.set_column(i, i, w)
            ws.freeze_panes(1,0)
        output.seek(0)
    except ModuleNotFoundError:
        from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
        from openpyxl.utils import get_column_letter
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            sheet="data"; out_df.to_excel(writer, index=False, sheet_name=sheet); ws=writer.sheets[sheet]
            header_fill = PatternFill("solid", fgColor="E2E8F0"); header_font=Font(bold=True)
            thin = Side(border_style="thin", color="CCCCCC"); border = Border(left=thin,right=thin,top=thin,bottom=thin)
            center = Alignment(horizontal="center", vertical="center", wrap_text=False)
            for cell in ws[1]:
                cell.fill=header_fill; cell.font=header_font; cell.alignment=center; cell.border=border; cell.number_format="@"
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.value not in (None,""):
                        cell.alignment=center; cell.border=border; cell.number_format="@"
            for col_idx, col_name in enumerate(out_df.columns, start=1):
                series = out_df[col_name].astype(str).tolist() if not out_df.empty else []
                width  = min(max([len(str(col_name))]+[len(s) for s in series]) + 2, 60) if series else len(str(col_name))
                ws.column_dimensions[get_column_letter(col_idx)].width = width
            ws.freeze_panes = "A2"
        output.seek(0)
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def _transform_fault_columns(df: pd.DataFrame) -> pd.DataFrame:
    """تحويل عمود 'نوع العطل' أو 'نوع الاعطال' إلى أعمدة منفصلة لكل نوع عطل مع القيم الافتراضية فارغة.''"""
    if df.empty or ('نوع العطل' not in df.columns and 'نوع الاعطال' not in df.columns):
        return df
    
    # إنشاء نسخة من DataFrame
    result_df = df.copy()
    
    # إضافة أعمدة الأعطال مع القيم الافتراضية فارغة
    for fault_type in FAULT_TYPES:
        result_df[fault_type] = ''
    
    # تحديث القيم بناءً على الأعطال المسجلة من كلا العمودين
    for idx, row in result_df.iterrows():
        raw1 = str(row.get('نوع العطل', '')).strip()
        raw2 = str(row.get('نوع الاعطال', '')).strip()
        merged = ",".join([s for s in [raw1, raw2] if s])
        if merged:
            faults = [f.strip() for f in merged.split(',') if f.strip()]
            for fault in faults:
                if fault in FAULT_TYPES:
                    result_df.at[idx, fault] = '1'
    
    # إزالة الأعمدة النصية للأعطال
    result_df = result_df.drop(columns=[c for c in ['نوع العطل','نوع الاعطال'] if c in result_df.columns])
    
    return result_df

def _project_recent_program_columns(df: pd.DataFrame) -> pd.DataFrame:
    """إسقاط وترتيب أعمدة المترددين (البيانات الحديثة) وفق المطلوب.
    الأعمدة النهائية:
    1- النوع (من 'القسم' إن لم يوجد)
    2- الادارة
    3- مكتب (من 'المكتب' إن لم يوجد)
    4- اسم العميل
    5- رقم العميل
    6- مسلسل
    7- التاريخ
    8- خدمات (من 'اسم المستخدم' إن وُجد)
    9- صيانه (من 'صيانة' أو 'القائم بالصيانة')
    10- الاذن (من 'رقم الإذن' أو بدائله)
    11- الأعطال (FAULT_TYPES بالترتيب)
    12- ملاحظات (من 'ملاحظات1' إن لم يوجد)
    13- _الفترة
    """
    if df is None or df.empty:
        return df

    out = df.copy()
    # تحويل أعمدة الأعطال
    out = _transform_fault_columns(out)

    # اشتقاق الأعمدة المطلوبة
    # النوع: من 'القسم' إن وُجد، وإلا تعبئة افتراضية باسم تبويب شاشة الاستعلام
    if 'النوع' not in out.columns:
        out['النوع'] = out.get('القسم', '') if 'القسم' in out.columns else out.get('التوع', '')
    try:
        out['النوع'] = out['النوع'].fillna('')
        out['النوع'] = out['النوع'].apply(lambda v: v if str(v).strip() else 'شاشة الاستعلام')
    except Exception:
        # في حال أي خطأ، نضمن وجود العمود بقيمة ثابتة
        out['النوع'] = 'شاشة الاستعلام'
    # مكتب
    if 'مكتب' not in out.columns and 'المكتب' in out.columns:
        out['مكتب'] = out['المكتب']
    # خدمات
    if 'خدمات' not in out.columns:
        out['خدمات'] = out.get('اسم المستخدم', '')
    # صيانه
    if 'صيانه' not in out.columns:
        if 'صيانة' in out.columns:
            out['صيانه'] = out['صيانة']
        else:
            out['صيانه'] = out.get('القائم بالصيانة', '')
    # رقم الإذن: دعم أسماء بديلة وإلا افتراضي فارغ
    if 'الاذن' not in out.columns:
        assigned = False
        for alt in ['رقم الإذن', 'رقم الاذن', 'الاذن', 'اذن']:
            if alt in out.columns:
                out['الاذن'] = out[alt]
                assigned = True
                break
        if not assigned:
            out['الاذن'] = ''
    # ملاحظات
    if 'ملاحظات' not in out.columns:
        out['ملاحظات'] = out.get('ملاحظات1', '') if 'ملاحظات1' in out.columns else ''
    # الحوالة المطلوبة
    if 'الحوالة المطلوبة' not in out.columns:
        out['الحوالة المطلوبة'] = ''
    # تأكيد وجود عمود الفترة
    if '_الفترة' not in out.columns:
        out['_الفترة'] = out.get('_الفترة', '')

    # تأكيد وجود أعمدة الأعطال المحددة فقط وبالترتيب
    for ft in FAULT_TYPES:
        if ft not in out.columns:
            out[ft] = ''

    desired = ['النوع', 'الادارة', 'مكتب', 'اسم العميل', 'رقم العميل', 'مسلسل', 'التاريخ', 'خدمات', 'صيانه', 'الاذن'] + FAULT_TYPES + ['الحوالة المطلوبة', 'ملاحظات', '_الفترة']

    # إسقاط الأعمدة الأخرى غير المطلوبة
    existing_desired = [c for c in desired if c in out.columns]
    projected = out[existing_desired]

    # إزالة عمود المصدر لمنع التكرار
    # نُلغي 'القسم', 'اسم المستخدم', 'القائم بالصيانة' إن كانت موجودة في النسخة الأصلية
    # لكننا لم نعدها ضمن projected لذا لسنا بحاجة لإسقاط جديد هنا
    return projected

def _years_list():
    rows = (ReportState.query
            .filter(ReportState.category.like("trader_frequent:%"))
            .order_by(ReportState.id.asc())
            .all())
    labels = []
    for r in rows:
        try: label = r.category.split(":", 1)[1]
        except Exception: continue
        # استبعاد المساعدات الداخلية من قائمة الفترات القديمة
        if label in {"__mapping__", "recent_program"}:
            continue
        if label not in labels: labels.append(label)
    return labels

@trader_services_bp.route("/")
@login_required
def index():
    return render_template("trader/index.html", sections=SECTIONS)


# ======================== Primary machines ========================
@trader_services_bp.route("/primary/import", methods=["GET"])
@login_required
@role_required("admin")
def primary_import_view():
    return render_template("trader/import_primary.html", title="استيراد الماكينات الأساسية للعملاء وماكينات الفرع")

@trader_services_bp.route("/primary/import", methods=["POST"])
@login_required
@role_required("admin")
def primary_import_post():
    f = request.files.get("file")
    if not f or not f.filename:
        flash("اختر ملف Excel.", "warning"); return redirect(url_for("trader_services_bp.primary_import_view"))
    df = _read_excel(f)
    _save_state("trader_primary", df=df)
    flash(f"تم استيراد الملف. سجلات: {len(df)}", "success")
    return redirect(url_for("trader_services_bp.primary_machines"))

@trader_services_bp.route("/primary", methods=["GET"])
@login_required
@permission_required('can_trader_primary')
def primary_machines():
    is_admin = getattr(current_user, "role", None) == "admin"
    row = _load_state("trader_primary")
    df = _json_to_df(row.data_json) if (row and row.data_json) else pd.DataFrame()

    map_row = _load_state("trader_primary:__mapping__")
    mapping = json.loads(map_row.mapping_json) if (map_row and map_row.mapping_json) else {}

    q = request.args.get("q","")
    search_in = request.args.get("search_in","all")
    page = request.args.get("page",1,type=int)
    page_size = request.args.get("page_size",25,type=int)
    page_size = 10 if page_size<10 else 1000 if page_size>1000 else page_size

    total=0; total_pages=1
    if not df.empty:
        mapped = _apply_mapping(df, mapping)
        filtered = _filter_dataframe(mapped, q, search_in)
        # إظهار كل الأعمدة طالما لم يبدأ البحث (q فارغ)، وعند البحث أسقط الأعمدة الفارغة
        visible_filtered = filtered if not (q and q.strip()) else _drop_empty_columns(filtered)
        page_df, total = _paginate(visible_filtered, page, page_size)
        total_pages = max(1, (total + page_size - 1)//page_size)
        search_cols = list(visible_filtered.columns)
    else:
        search_cols = []
        page_df = pd.DataFrame()

    cols = list(page_df.columns)[:50]
    rows = page_df[cols].to_dict(orient="records") if not page_df.empty else []

    def _page_url(n): 
        return url_for("trader_services_bp.primary_machines", q=q, search_in=search_in, page=n, page_size=page_size)
    first_pages = [n for n in [1,2,3] if n <= total_pages]
    pagination = {
        "prev": _page_url(page-1) if page>1 else None,
        "next": _page_url(page+1) if page<total_pages else None,
        "first_pages": [{"n":n,"url":_page_url(n),"active":(n==page)} for n in first_pages],
        "show_ellipsis": total_pages>3,
        "last": {"n":total_pages,"url":_page_url(total_pages),"active":(page==total_pages)} if total_pages>3 else None,
        "page": page, "total_pages": total_pages
    }

    return render_template("trader/primary.html",
                           title="الماكينات الأساسية للعملاء وماكينات الفرع",
                           is_admin=is_admin, has_data=not df.empty,
                           q=q, search_in=search_in,
                           page=page, page_size=page_size,
                           cols=cols, rows=rows,
                           order_csv=",".join(mapping.get("order", [])),
                           rename_lines="\n".join([f"{k}=>{v}" for k,v in (mapping.get("rename") or {}).items()]),
                           pagination=pagination,
                           search_cols=search_cols)

@trader_services_bp.route("/primary/save_mapping", methods=["POST"])
@login_required
@role_required("admin")
def primary_save_mapping():
    order_csv = (request.form.get("order_csv") or "").strip()
    order = [c.strip() for c in order_csv.split(",") if c.strip()] if order_csv else []
    rename_lines = (request.form.get("rename_lines") or "").strip()
    rename = {}
    if rename_lines:
        for line in rename_lines.splitlines():
            if "=>" in line:
                old, new = line.split("=>",1)
                old, new = old.strip(), new.strip()
                if old and new: rename[old] = new
    _save_state("trader_primary:__mapping__", mapping={"order":order, "rename":rename})
    return redirect(url_for("trader_services_bp.primary_machines"))

@trader_services_bp.route("/primary/export", methods=["GET"])
@login_required
@permission_required('can_trader_primary')
def primary_export():
    row = _load_state("trader_primary")
    if not row or not row.data_json:
        flash("لا توجد بيانات لتصديرها.", "warning")
        return redirect(url_for("trader_services_bp.primary_machines"))
    map_row = _load_state("trader_primary:__mapping__")
    mapping = json.loads(map_row.mapping_json) if (map_row and map_row.mapping_json) else {}
    q = request.args.get("q","")
    search_in = request.args.get("search_in","all")
    out = _apply_mapping(_json_to_df(row.data_json), mapping)
    out = _filter_dataframe(out, q, search_in)
    out = _drop_empty_columns(out)
    out = _coerce_all_text_no_decimals(out)
    if out.empty:
        flash("لا توجد بيانات لتصديرها.", "warning")
        return redirect(url_for("trader_services_bp.primary_machines"))
    return _excel_response(out, "الماكينات_الأساسية_وماكينات_الفرع.xlsx")


# ======================== Frequent visitors (new screen) ========================
# عرض شاشة المترددين: بيانات حديثة فقط
@trader_services_bp.route("/frequent", methods=["GET"])
@login_required
@permission_required('can_trader_frequent')
def frequent_visitors():
    # تم إلغاء تبويب البيانات القديمة — العرض حديث فقط
    tab = "recent"
    q = request.args.get("q", "")
    search_in = request.args.get("search_in", "all")
    page = request.args.get("page", 1, type=int)
    page_size = request.args.get("page_size", 25, type=int)
    page_size = 10 if page_size < 10 else 1000 if page_size > 1000 else page_size
    label = request.args.get("label", "")

    is_admin = getattr(current_user, "role", None) == "admin"
    years = []

    df = pd.DataFrame(); mapping = {}
    has_data = False; search_cols = []; cols = []; rows = []
    total = 0; total_pages = 1

    # بيانات حديثة من برنامج الإضافة الحديث
    row = _load_state("trader_frequent:recent_program")
    map_row = _load_state("trader_frequent:__mapping__")
    mapping = json.loads(map_row.mapping_json) if (map_row and map_row.mapping_json) else {}
    df = _json_to_df(row.data_json) if (row and row.data_json) else pd.DataFrame()
    df = _project_recent_program_columns(df)
    df = _apply_mapping(df, mapping)
    has_data = not df.empty

    if has_data:
        filtered = _filter_dataframe(df, q, search_in)
        # ضمان الأعمدة المطلوبة في تبويب البيانات الحديثة قبل إسقاط الأعمدة الفارغة
        if tab == "recent":
            # النوع: نحاول من 'النوع' أو 'القسم' أو 'التوع'
            if 'النوع' not in filtered.columns:
                _src = next((c for c in ['النوع', 'القسم', 'التوع'] if c in filtered.columns), None)
                filtered['النوع'] = filtered[_src] if _src else ''
            # المكتب: من 'مكتب' أو 'المكتب'
            if 'مكتب' not in filtered.columns:
                _src = next((c for c in ['مكتب', 'المكتب'] if c in filtered.columns), None)
                filtered['مكتب'] = filtered[_src] if _src else ''
            # رقم الإذن: نثبت عمود 'الاذن' بقراءة أي بديل موجود
            if 'الاذن' not in filtered.columns:
                _src = next((c for c in ['الاذن', 'رقم الإذن', 'رقم الاذن', 'اذن'] if c in filtered.columns), None)
                filtered['الاذن'] = filtered[_src] if _src else ''

        # إظهار كل الأعمدة طالما لم يبدأ البحث (q فارغ)، وعند البحث أسقط الأعمدة الفارغة
        visible_filtered = filtered if not (q and q.strip()) else _drop_empty_columns(filtered)
        # إعادة ترتيب الأعمدة لإبراز الأعمدة الأساسية في الحديثة
        if not visible_filtered.empty:
            _order = DEFAULT_FREQUENT_ORDER
            _ordered_cols = [c for c in _order if c in visible_filtered.columns]
            _ordered_cols += [c for c in visible_filtered.columns if c not in _ordered_cols]
            visible_filtered = visible_filtered[_ordered_cols]
        page_df, total = _paginate(visible_filtered, page, page_size)
        total_pages = max(1, (total + page_size - 1)//page_size)
        # تقييد أعمدة البحث للقائمة المطلوبة مع دعم بعض المسميات البديلة
        present = set(visible_filtered.columns)
        desired_base = ['النوع','الادارة','اسم العميل','رقم العميل','مسلسل','التاريخ','خدمات']
        search_cols = [c for c in desired_base if c in present]
        # مجموعات بدائل تعرض أول اسم موجود منها
        for group in [
            ['مكتب','المكتب'],
            ['الاذن','رقم الإذن','رقم الاذن','اذن'],
            ['صيانه','صيانة']
        ]:
            found = next((c for c in group if c in present), None)
            if found and found not in search_cols:
                search_cols.append(found)
        # استبعاد الأعمدة غير المرغوبة صراحةً من قائمة البحث (تطبيعًا)
        search_cols = [c for c in search_cols if _normalize_name(c) not in EXCLUDED_SEARCH_COLUMNS_NORM]
        cols = list(page_df.columns)[:50]
        # تضمين فهرس السطر الأصلي لتمكين الحذف بدون الاعتماد على أي معرّفات
        rows = []
        if not page_df.empty:
            for _idx, _row in page_df[cols].iterrows():
                d = _row.to_dict()
                d['__index'] = int(_idx)
                rows.append(d)

    def _page_url(n):
        return url_for("trader_services_bp.frequent_visitors", tab="recent", q=q, search_in=search_in, page=n, page_size=page_size, label=label)
    first_pages = [n for n in [1,2,3] if n <= total_pages]
    pagination = {
        "prev": _page_url(page-1) if page>1 else None,
        "next": _page_url(page+1) if page<total_pages else None,
        "first_pages": [{"n":n,"url":_page_url(n),"active":(n==page)} for n in first_pages],
        "show_ellipsis": total_pages>3,
        "last": {"n":total_pages,"url":_page_url(total_pages),"active":(page==total_pages)} if total_pages>3 else None,
        "page": page, "total_pages": total_pages
    }

    # أسماء القائمين بالصيانة لعرضها في قائمة منسدلة بالتعديل (مطلوبة تحديدًا)
    maintenance_names = ["حمدي", "اسلام", "امين", "مصطفى", "وائل"]

    return render_template("trader/frequent.html",
                           title="المترددين",
                           is_admin=is_admin, has_data=has_data,
                           tab=tab, years=years, current_label=(label or ""),
                           q=q, search_in=search_in,
                           page=page, page_size=page_size,
                           cols=cols, rows=rows,
                           search_cols=search_cols,
                           pagination=pagination,
                           fault_types=FAULT_TYPES,
                           maintenance_names=maintenance_names)


# شاشة الاستيراد — حفظ المابينج
@trader_services_bp.route("/frequent/import", methods=["GET"])
@login_required
@role_required("admin")
def frequent_import_view():
    map_row = _load_state("trader_frequent:__mapping__")
    mapping = json.loads(map_row.mapping_json) if (map_row and map_row.mapping_json) else {}
    order_csv = ",".join(mapping.get("order", []))
    rename_lines = "\n".join([f"{k}=>{v}" for k,v in (mapping.get("rename") or {}).items()])
    return render_template("trader/import_frequent.html", title="استيراد المترددين", order_csv=order_csv, rename_lines=rename_lines)


@trader_services_bp.route("/frequent/import", methods=["POST"])
@login_required
@role_required("admin")
def frequent_import_post():
    label = (request.form.get("label") or "").strip()
    if not label:
        flash("أدخل اسم الفترة/الملف للاستيراد.", "warning"); return redirect(url_for("trader_services_bp.frequent_import_view"))
    f1 = request.files.get("file1")
    f2 = request.files.get("file2")
    if (not f1 or not f1.filename) and (not f2 or not f2.filename):
        flash("اختر ملفًا واحدًا على الأقل للاستيراد.", "warning"); return redirect(url_for("trader_services_bp.frequent_import_view"))

    # قراءة الملفات ودمجها
    dfs = []
    try:
        if f1 and f1.filename: dfs.append(_read_excel(f1))
        if f2 and f2.filename: dfs.append(_read_excel(f2))
    except Exception as ex:
        flash(f"تعذر قراءة ملفات Excel: {ex}", "danger"); return redirect(url_for("trader_services_bp.frequent_import_view"))
    df = pd.concat(dfs, ignore_index=True) if len(dfs) > 1 else (dfs[0] if dfs else pd.DataFrame())

    # حفظ المابينج (اختياري)
    order_csv = (request.form.get("order_csv") or "").strip()
    rename_lines = (request.form.get("rename_lines") or "").strip()
    mapping = {}
    if order_csv:
        mapping["order"] = [c.strip() for c in order_csv.split(",") if c.strip()]
    if rename_lines:
        rename = {}
        for line in rename_lines.splitlines():
            if "=>" in line:
                src, dst = [s.strip() for s in line.split("=>", 1)]
                if src and dst: rename[src] = dst
        if rename: mapping["rename"] = rename
    if mapping:
        _save_state("trader_frequent:__mapping__", mapping=mapping)

    # حفظ البيانات تحت الفترة المعطاة
    _save_state(f"trader_frequent:{label}", df=df)
    flash(f"تم الاستيراد بنجاح للفترة: {label}. سجلات: {len(df)}", "success")
    return redirect(url_for("trader_services_bp.frequent_visitors", tab="recent", label=label))


@trader_services_bp.route("/frequent/save_mapping", methods=["POST"])
@login_required
@role_required("admin")
def frequent_save_mapping():
    order_csv = (request.form.get("order_csv") or "").strip()
    order = [c.strip() for c in order_csv.split(",") if c.strip()] if order_csv else []
    rename_lines = (request.form.get("rename_lines") or "").strip()
    rename = {}
    if rename_lines:
        for line in rename_lines.splitlines():
            if "=>" in line:
                old, new = line.split("=>",1)
                old, new = old.strip(), new.strip()
                if old and new: rename[old] = new
    _save_state("trader_frequent:__mapping__", mapping={"order":order, "rename":rename})
    return redirect(url_for("trader_services_bp.frequent_visitors", tab="recent"))


@trader_services_bp.route("/frequent/export", methods=["GET"])
@login_required
@permission_required('can_trader_frequent')
def frequent_export():
    tab = "recent"
    q = request.args.get("q", "")
    search_in = request.args.get("search_in", "all")
    label = request.args.get("label", "")

    row = _load_state("trader_frequent:recent_program")
    if not row or not row.data_json:
        flash("لا توجد بيانات لتصديرها.", "warning"); return redirect(url_for("trader_services_bp.frequent_visitors", tab="recent"))
    map_row = _load_state("trader_frequent:__mapping__")
    mapping = json.loads(map_row.mapping_json) if (map_row and map_row.mapping_json) else {}
    out = _project_recent_program_columns(_json_to_df(row.data_json))
    out = _apply_mapping(out, mapping)

    out = _filter_dataframe(out, q, search_in)
    out = _drop_empty_columns(out)
    out = _coerce_all_text_no_decimals(out)
    if out.empty:
        flash("لا توجد بيانات لتصديرها.", "warning")
        return redirect(url_for("trader_services_bp.frequent_visitors", tab=tab, label=label))
    return _excel_response(out, "المترددين.xlsx")


# تحديث سجل حديث (صاحب السجل أو الأدمن فقط) — يمنع تعديل التاريخ
@trader_services_bp.route("/frequent/update_recent", methods=["POST"])
@login_required
@permission_required('can_trader_frequent')
def frequent_update_recent():
    try:
        payload = request.get_json(silent=True) or {}
        serial = (payload.get('serial') or '').strip()
        order_number = (payload.get('order_number') or '').strip()
        date = (payload.get('date') or '').strip()
        row_index_raw = payload.get('row_index')
        updates = payload.get('updates') or {}
        # نسمح بالتعريف إما بفهرس الصف أو بالحقول التعريفية
        if (row_index_raw is None or str(row_index_raw).strip() == "") and (not serial or not order_number or not date):
            return jsonify({"success": False, "message": "مطلوب فهرس الصف أو حقول التعريف (مسلسل + الإذن + التاريخ)."}), 400

        row = _load_state("trader_frequent:recent_program")
        df = _json_to_df(row.data_json) if (row and row.data_json) else pd.DataFrame()
        if df.empty:
            return jsonify({"success": False, "message": "لا توجد بيانات للتحديث."}), 404

        # تحديد الصف المستهدف: أولاً بفهرس الصف إن وُجد، وإلا عبر التعريف القديم
        idx = None
        if row_index_raw is not None and str(row_index_raw).strip() != "":
            try:
                _idx = int(str(row_index_raw).strip())
            except ValueError:
                return jsonify({"success": False, "message": "فهرس الصف غير صالح."}), 400
            if _idx in df.index:
                idx = _idx
            else:
                return jsonify({"success": False, "message": "لم يتم العثور على السجل (فهرس غير موجود)."}), 404
        else:
            mask = (df.get('مسلسل', '').astype(str) == serial) & (df.get('الاذن', '').astype(str) == order_number) & (df.get('التاريخ', '').astype(str) == date)
            idx_list = df.index[mask].tolist()
            if not idx_list:
                return jsonify({"success": False, "message": "لم يتم العثور على السجل المستهدف."}), 404
            idx = idx_list[0]

        # تحقق الصلاحية: اسم المستخدم صاحب السجل أو الأدمن
        owner = str(df.at[idx, 'اسم المستخدم']) if 'اسم المستخدم' in df.columns else ''
        is_admin = getattr(current_user, 'role', None) == 'admin'
        if not is_admin and owner != getattr(current_user, 'username', ''):
            return jsonify({"success": False, "message": "غير مسموح بتعديل هذا السجل."}), 403

        # منع تعديل التاريخ وأي حقول غير مسموح بها
        if 'التاريخ' in updates:
            updates.pop('التاريخ', None)

        # الحقول المسموحة بالتعديل فقط
        allowed_non_fault = {'صيانه', 'صيانة', 'الاذن', 'ملاحظات', 'الحوالة المطلوبة'}
        fault_updates = {k: v for k, v in updates.items() if k in FAULT_TYPES}
        safe_updates = {k: v for k, v in updates.items() if (k in allowed_non_fault)}

        # السماح بإضافة أعطال جديدة بإنشاء الأعمدة الناقصة تلقائيًا
        for fault_name, val in fault_updates.items():
            if fault_name not in df.columns:
                df[fault_name] = ''
            df.at[idx, fault_name] = '1' if str(val).strip() == '1' else ''

        # تطبيق التحديثات المسموحة فقط
        for k, v in safe_updates.items():
            # دعم صيغتي "صيانه" و"صيانة"
            if k in {'صيانه', 'صيانة'}:
                target_col = 'صيانه' if 'صيانه' in df.columns else ('صيانة' if 'صيانة' in df.columns else 'صيانه')
                if target_col not in df.columns:
                    df[target_col] = ''
                df.at[idx, target_col] = _textify(v)
                continue
            # باقي الحقول المسموحة
            if k not in df.columns:
                df[k] = ''
            df.at[idx, k] = _textify(v)

        row.data_json = _df_to_json(df); db.session.commit()
        return jsonify({"success": True, "message": "تم تحديث السجل."})
    except Exception as ex:
        current_app.logger.exception("frequent_update_recent error:")
        db.session.rollback()
        return jsonify({"success": False, "message": f"حدث خطأ أثناء التحديث: {ex}"}), 500


# حذف سجل حديث (صاحب السجل أو الأدمن فقط)
@trader_services_bp.route("/frequent/delete_recent_row", methods=["POST"])
@login_required
@permission_required('can_trader_frequent')
def frequent_delete_recent_row():
    try:
        payload = request.get_json(silent=True) or {}
        row = _load_state("trader_frequent:recent_program")
        df = _json_to_df(row.data_json) if (row and row.data_json) else pd.DataFrame()
        if df.empty:
            return jsonify({"success": False, "message": "لا توجد بيانات للحذف."}), 404

        # دعم الحذف بفهرس الصف مباشرة
        row_index_raw = payload.get('row_index')
        if row_index_raw is not None and str(row_index_raw).strip() != "":
            try:
                idx = int(str(row_index_raw).strip())
            except ValueError:
                return jsonify({"success": False, "message": "فهرس الصف غير صالح."}), 400
            if idx not in df.index:
                return jsonify({"success": False, "message": "لم يتم العثور على السجل (فهرس غير موجود)."}), 404
            # تحقق الصلاحية: المستخدم صاحب السجل أو أدمن
            is_admin = getattr(current_user, 'role', None) == 'admin'
            if not is_admin:
                owner = str(df.at[idx, 'خدمات']) if 'خدمات' in df.columns else str(df.at[idx, 'اسم المستخدم']) if 'اسم المستخدم' in df.columns else ''
                if owner != getattr(current_user, 'username', ''):
                    return jsonify({"success": False, "message": "غير مسموح بحذف سجلات لا تملكها."}), 403
            df = df.drop(index=[idx]).reset_index(drop=True)
            row.data_json = _df_to_json(df); db.session.commit()
            return jsonify({"success": True, "message": "تم حذف السجل."})

        # خلاف ذلك استمرار دعم الحذف بالمعرّفات المتوفرة للحفاظ على التوافق
        serial = (payload.get('serial') or '').strip()
        order_number = (payload.get('order_number') or '').strip()
        date = (payload.get('date') or '').strip()
        if not (serial or order_number or date):
            return jsonify({"success": False, "message": "يرجى تحديد فهرس الصف أو تقديم المسلسل/رقم الإذن/التاريخ."}), 400

        # بناء قناع اعتمادًا على الحقول المتوفرة فقط
        mask_parts = []
        if serial:
            mask_parts.append(df.get('مسلسل', '').astype(str) == serial)
        if order_number:
            # دعم أسماء بديلة محتملة داخل البيانات
            if 'الاذن' in df.columns:
                mask_parts.append(df.get('الاذن', '').astype(str) == order_number)
            else:
                # في حال كانت البيانات قبل الإسقاط النهائي
                for alt in ['رقم الإذن', 'رقم الاذن', 'اذن']:
                    if alt in df.columns:
                        mask_parts.append(df.get(alt, '').astype(str) == order_number)
                        break
        if date:
            mask_parts.append(df.get('التاريخ', '').astype(str) == date)

        mask = mask_parts[0] if len(mask_parts)==1 else mask_parts[0]
        for part in mask_parts[1:]:
            mask = mask & part
        idx_list = df.index[mask].tolist()
        if not idx_list:
            return jsonify({"success": False, "message": "لم يتم العثور على أي سجلات مطابقة للحذف."}), 404
        if len(idx_list) > 1:
            return jsonify({"success": False, "message": "المعرفات المقدمة تطابق أكثر من سجل. يرجى إضافة بيان تعريف آخر لتحديد السجل بدقة."}), 409

        # تحقق الصلاحية: المستخدم صاحب السجل أو أدمن
        is_admin = getattr(current_user, 'role', None) == 'admin'
        if not is_admin:
            idx = idx_list[0]
            owner = str(df.at[idx, 'اسم المستخدم']) if 'اسم المستخدم' in df.columns else ''
            if owner != getattr(current_user, 'username', ''):
                return jsonify({"success": False, "message": "غير مسموح بحذف سجلات لا تملكها."}), 403

        df = df.drop(index=idx_list).reset_index(drop=True)
        row.data_json = _df_to_json(df); db.session.commit()
        return jsonify({"success": True, "message": "تم حذف السجل."})
    except Exception as ex:
        current_app.logger.exception("frequent_delete_recent_row error:")
        db.session.rollback()
        return jsonify({"success": False, "message": f"حدث خطأ أثناء الحذف: {ex}"}), 500


# إضافة سجل جديد إلى بيانات المترددين الحديثة (من شاشة الاستعلام)
@trader_services_bp.route("/frequent/add_recent", methods=["POST"])
@login_required
@permission_required('can_inquiry')
def frequent_add_recent():
    try:
        payload = request.get_json(silent=True) or {}
        serial = (payload.get('machine_serial') or payload.get('serial') or '').strip()
        machine_code = (payload.get('machine_code') or '').strip()
        order_number = (payload.get('order_number') or '').strip()
        required_transfer = (payload.get('required_transfer') or '').strip()
        maintenance = (payload.get('maintenance') or '').strip()
        # حقول إضافية مطلوبة من شاشة الاستعلام
        management = (payload.get('management') or payload.get('department') or '').strip()
        office = (payload.get('office') or payload.get('branch_office') or '').strip()
        client_name = (payload.get('client_name') or '').strip()
        client_number = (payload.get('client_number') or '').strip()
        section = (payload.get('section') or payload.get('القسم') or '').strip()
        # التاريخ يُحدد تلقائيًا ولا يُسمح بإرساله من العميل
        from datetime import datetime
        # حفظ التاريخ مع الوقت لضمان عرض وقت الزيارة في شاشة الاستعلام
        date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        notes = (payload.get('notes') or '').strip()
        fault_types = payload.get('fault_types') or []

        if not serial:
            return jsonify({"success": False, "message": "المسلسل مطلوب لإضافة السجل."}), 400

        # فرض المتطلبات الأساسية: على الأقل عطل واحد + رقم إذن أرقام فقط
        if not (isinstance(fault_types, list) and len([ft for ft in fault_types if str(ft).strip()]) > 0):
            return jsonify({"success": False, "message": "يجب اختيار عطل واحد على الأقل."}), 400
        if not (order_number and order_number.isdigit()):
            return jsonify({"success": False, "message": "رقم الإذن مطلوب ويجب أن يكون أرقام فقط."}), 400

        # جلب (أو إنشاء) أحدث سجل recent_program المرتبط بالأدمن إن وجد
        from models import User
        admin_user = User.query.filter_by(role='admin').first()
        row = (ReportState.query
               .filter(ReportState.category == 'trader_frequent:recent_program', ReportState.user_id == (admin_user.id if admin_user else None))
               .order_by(ReportState.created_at.desc())
               .first())
        if not row:
            # إنشاء سجل فارغ في حال عدم وجوده
            row = ReportState(category='trader_frequent:recent_program', user_id=(admin_user.id if admin_user else getattr(current_user, 'id', None)))
            db.session.add(row)
            row.data_json = _df_to_json(pd.DataFrame())

        df = _json_to_df(row.data_json)
        # بناء صف البيانات الجديد
        new_rec = {
            'الادارة': management,
            'مكتب': office,
            'اسم العميل': client_name,
            'رقم العميل': client_number,
            'القسم': section,
            'مسلسل': serial,
            'رقم الماكينة': machine_code,
            'التاريخ': date,
            'خدمات': getattr(current_user, 'username', ''),
            'اسم المستخدم': getattr(current_user, 'username', ''),
            'صيانه': maintenance,
            'الاذن': order_number,
            'الحوالة المطلوبة': required_transfer,
            'ملاحظات': notes,
        }
        # وضع أعطال مختارة على شكل أعمدة بقيمة '1'
        if isinstance(fault_types, list):
            for ft in fault_types:
                ft = (str(ft) or '').strip()
                if not ft:
                    continue
                if ft not in df.columns:
                    # تأكد من وجود العمود في الإطار
                    df[ft] = ""
                new_rec[ft] = '1'

        # إضافة الصف
        try:
            df = pd.concat([df, pd.DataFrame([new_rec])], ignore_index=True) if not df.empty else pd.DataFrame([new_rec])
        except Exception:
            # fallback آمن إذا فشل concat بسبب أنواع الأعمدة
            df = pd.DataFrame([new_rec]) if df.empty else df.append(new_rec, ignore_index=True)

        row.data_json = _df_to_json(df)
        db.session.commit()
        return jsonify({"success": True, "message": "تمت إضافة السجل بنجاح."})
    except Exception as ex:
        current_app.logger.exception("frequent_add_recent error:")
        db.session.rollback()
        return jsonify({"success": False, "message": f"حدث خطأ أثناء الإضافة: {ex}"}), 500


# ======================== Recent frequent count API ========================
@trader_services_bp.route("/frequent/recent_count", methods=["GET"])
@login_required
@permission_required('can_inquiry')
def frequent_recent_count():
    """إرجاع عدد سجلات قسم المترددين الحديثة + تفاصيل المسلسلات إن وُجدت."""
    try:
        # قراءة أحدث بيانات recent_program (مع تطبيق الـ mapping إن وُجد)
        row = _load_state("trader_frequent:recent_program")
        df = _json_to_df(row.data_json) if (row and row.data_json) else pd.DataFrame()
        map_row = _load_state("trader_frequent:__mapping__")
        mapping = json.loads(map_row.mapping_json) if (map_row and map_row.mapping_json) else {}
        if not df.empty:
            df = _apply_mapping(df, mapping)
            df = _drop_empty_columns(df)
            # إسقاط/ترتيب الأعمدة وفق نموذج الحديثة لضمان توفر الأسماء الموحدة
            df = _project_recent_program_columns(df)

        total = int(df.shape[0]) if not df.empty else 0
        details = {}
        if not df.empty and ('مسلسل' in df.columns):
            try:
                details = df.groupby('مسلسل').size().to_dict()
            except Exception:
                details = {}

        return jsonify({
            "success": True,
            "label": "البيانات الحديثة (البرنامج)",
            "total": total,
            "serial_details": details
        })
    except Exception as ex:
        current_app.logger.exception("frequent_recent_count error:")
        return jsonify({"success": False, "message": f"خطأ أثناء حساب العدد: {ex}"}), 500
