from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_login import login_required, current_user
from models import db, SupportCase, User
import pandas as pd
import io, re, os
from decimal import Decimal, InvalidOperation
from datetime import datetime, timedelta

support_bp = Blueprint("support_bp", __name__, url_prefix="/support")

# ===== إعداد المنطقة الزمنية (اكتشاف تلقائي مع إمكانية التخصيص عبر البيئة) =====
def _get_local_utc_offset_minutes() -> int:
    """يحسب فرق التوقيت المحلي عن UTC بالدقائق.
    يستخدم الفرق بين datetime.now() و datetime.utcnow().
    """
    try:
        delta = datetime.now() - datetime.utcnow()
        # تقريب إلى أقرب دقيقة لتجنب انحرافات الثواني
        return int(round(delta.total_seconds() / 60.0))
    except Exception:
        # افتراضي منطقي إذا فشل الحساب
        return 0

_TZ_OFFSET_MIN = _get_local_utc_offset_minutes()
env_tz = os.environ.get("TIMEZONE_OFFSET_MINUTES")
if env_tz and re.fullmatch(r"-?\d+", env_tz.strip()):
    try:
        _TZ_OFFSET_MIN = int(env_tz.strip())
    except Exception:
        pass

def _to_utc_str_from_local(dt_str: str | None) -> str:
    """حوّل datetime-local المدخَل محليًا إلى UTC string 'YYYY-MM-DD HH:MM'."""
    if not dt_str:
        return ""
    s = dt_str.strip()
    for fmt in ("%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            dt_local = datetime.strptime(s, fmt)
            break
        except Exception:
            continue
    else:
        return ""
    # اطرح الإزاحة المحلية للحصول على UTC (يعتمد على الإعداد/الاكتشاف)
    dt_utc = dt_local - timedelta(minutes=_TZ_OFFSET_MIN)
    return dt_utc.strftime("%Y-%m-%d %H:%M")

def _parse_utc_str(dt_or_str):
    """حوّل قيمة زمنية إلى datetime (UTC naive).
    يقبل إما نصًا بالتنسيق 'YYYY-MM-DD HH:MM' أو كائن datetime جاهز.
    """
    if not dt_or_str:
        return None
    if isinstance(dt_or_str, datetime):
        return dt_or_str
    s = str(dt_or_str).strip().replace("T", " ")
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            continue
    return None

def _utc_now():
    return datetime.utcnow()

def _utc_now_floor_minute():
    """ارجع الوقت الحالي UTC مع تصفير الثواني والميكروثواني (بداية الدقيقة)."""
    now = datetime.utcnow()
    return now.replace(second=0, microsecond=0)

def _to_local_display(utc_dt_or_str):
    """تحول التوقيت المحفوظ (UTC) إلى التوقيت المحلي (UTC+3) مع الوقت."""
    utc_dt = utc_dt_or_str
    if isinstance(utc_dt, str):
        utc_dt = _parse_utc_str(utc_dt)
    if not utc_dt: return ""
    
    # التطبيق الصحيح للإزاحة للعرض
    local_dt = utc_dt + timedelta(minutes=_TZ_OFFSET_MIN)
    # التأكد من أن صيغة العرض تشمل التاريخ والوقت والدقائق
    return local_dt.strftime("%Y-%m-%d %H:%M")

# رموز تُعتبر "قيمة فارغة" لنعومتها من العرض/التصدير
_EMPTY_TOKENS = {"nan", "none", "null", "na", "n/a", "nat", "-", "—"}

def _textify(v) -> str:
    # 💥 استثناء خاص لـ "وقت التذكير" لضمان ظهوره كنص كامل
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M") 
    t = str(v).strip()
    if t == "" or t.lower() in _EMPTY_TOKENS:
        return ""
    t = t.replace(",", "")
    if re.fullmatch(r"\d+", t):
        return t
    if re.fullmatch(r"\d+\.\d+", t):
        # تم الإبقاء على هذا المنطق لحذف الأجزاء العشرية (كما كان)
        return t.split(".", 1)[0]
    if re.fullmatch(r"[0-9]+(\.[0-9]+)?[eE][+\-]?[0-9]+", t):
        try:
            # تم الإبقاء على هذا المنطق لحذف الأجزاء العشرية (كما كان)
            return str(Decimal(t).to_integral_value(rounding="ROUND_DOWN"))
        except InvalidOperation:
            return ""
    return t

def _df_text(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    for c in out.columns:
        out[c] = out[c].map(_textify)
    out.columns = [str(c).strip().replace("\n"," ").replace("\r"," ") for c in out.columns]
    return out

def _drop_empty_columns(df: pd.DataFrame) -> pd.DataFrame:
    """إزالة الأعمدة التي لا تحتوي إلا على قيم فارغة أو رموز تعتبر فارغة."""
    if df is None or df.empty:
        return df

    def _is_empty_series(s: pd.Series) -> bool:
        # نحول القيم إلى نصوص لضمان تطبيق التحقق على رموز "الفارغ" النصية
        vals = s.fillna("").astype(str).str.strip().str.lower()
        return ((vals == "") | (vals.isin(_EMPTY_TOKENS))).all()
    
    # تحديد الأعمدة التي يجب الإبقاء عليها (غير الفارغة بالكامل)
    cols_to_keep = [c for c in df.columns if not _is_empty_series(df[c])]

    return df[cols_to_keep]

def _filter_dataframe(df: pd.DataFrame, q: str, search_in: str | None) -> pd.DataFrame:
    if not q:
        return df
    ql = q.strip().lower()
    if not ql:
        return df
    if not search_in or search_in == "all" or search_in not in df.columns:
        mask = df.apply(lambda col: col.astype(str).str.lower().str.contains(ql, na=False))
        return df[mask.any(axis=1)]
    return df[df[search_in].astype(str).str.lower().str.contains(ql, na=False)]

def _paginate(df: pd.DataFrame, page: int, page_size: int):
    n = len(df)
    start = max(0, (page - 1) * page_size)
    end = start + page_size
    return df.iloc[start:end], n

def _excel_response(out_df: pd.DataFrame, filename: str):
    output = io.BytesIO()
    try:
        import xlsxwriter
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            sheet = "data"
            out_df.to_excel(writer, index=False, sheet_name=sheet)
            ws = writer.sheets[sheet]
            book = writer.book
            
            # تنسيق التاريخ والوقت في Excel
            # **يجب التأكد من أن الأعمدة التي تحتوي على التاريخ هي من نوع datetime قبل تصديرها هنا
            date_time_format = book.add_format({'num_format': 'yyyy-mm-dd hh:mm', 'align': 'center', "valign": "vcenter", "border": 1})
            
            header_fmt = book.add_format({"bold": True, "bg_color": "#E2E8F0", "align": "center", "valign": "vcenter", "border": 1, "num_format": "@"})
            cell_fmt = book.add_format({"align": "center", "valign": "vcenter", "border": 1, "num_format": "@"})
            
            for col_idx, col_name in enumerate(out_df.columns):
                ws.write(0, col_idx, col_name, header_fmt)
                
                # تطبيق تنسيق التاريخ والوقت على الأعمدة المطلوبة
                if col_name in ["وقت التذكير", "تاريخ التسجيل"]:
                    ws.set_column(col_idx, col_idx, 20, date_time_format)
                else:
                    # تحديد عرض افتراضي مع تطبيق تنسيق نص عام بشكل صحيح
                    ws.set_column(col_idx, col_idx, 20, cell_fmt)
            
            # تم إزالة التنسيق الشرطي العام الذي كان يفرض صيغة النص على كل الخلايا
            # حتى لا يطغى على تنسيق التاريخ/الوقت للأعمدة الزمنية
            
            # تعديل عرض الأعمدة لتناسب المحتوى
            for i, c in enumerate(out_df.columns):
                series = out_df[c].astype(str)
                w = min(max([len(str(c))] + [len(s) for s in series.tolist()]) + 2, 60)
                ws.set_column(i, i, w)
                
            ws.freeze_panes(1, 0)
        output.seek(0)
    except ModuleNotFoundError:
        # (فرع openpyxl) تطبيق تنسيق دقيق للتواريخ وعدم تحويلها إلى نص
        from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
        from openpyxl.utils import get_column_letter
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            sheet = "data"
            out_df.to_excel(writer, index=False, sheet_name=sheet)
            ws = writer.sheets[sheet]

            header_fill = PatternFill("solid", fgColor="E2E8F0")
            header_font = Font(bold=True)
            thin = Side(border_style="thin", color="CCCCCC")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            center = Alignment(horizontal="center", vertical="center", wrap_text=False)

            # الأعمدة الزمنية المطلوب تطبيق تنسيق التاريخ عليها
            date_cols = {"وقت التذكير", "تاريخ التسجيل"}
            col_name_by_idx = {idx+1: name for idx, name in enumerate(out_df.columns)}

            # تنسيق رأس الجدول
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border
                # الرأس دائمًا نص
                cell.number_format = "@"

            # تنسيق خلايا البيانات مع تمييز أعمدة التاريخ
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=2):
                for col_idx, cell in enumerate(row, start=1):
                    if cell.value in (None, ""):
                        continue
                    cell.alignment = center
                    cell.border = border
                    col_name = col_name_by_idx.get(col_idx)
                    if col_name in date_cols:
                        # تنسيق تاريخ/وقت قياسي: yyyy-mm-dd hh:mm
                        cell.number_format = "yyyy-mm-dd hh:mm"
                    else:
                        cell.number_format = "@"

            # ضبط العرض لكل عمود وفق المحتوى
            for col_idx, col_name in enumerate(out_df.columns, start=1):
                series = out_df[col_name].astype(str).tolist() if not out_df.empty else []
                width = min(max([len(str(col_name))] + [len(s) for s in series]) + 2, 60) if series else len(str(col_name))
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            ws.freeze_panes = "A2"
        output.seek(0)
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ========== عرض/بحث/تصدير ==========
@support_bp.route("/", methods=["GET"])
@login_required
def index():
    q = request.args.get("q", "")
    search_in = request.args.get("search_in", "all")
    page = request.args.get("page", 1, type=int)
    page_size = request.args.get("page_size", 25, type=int)
    page_size = 10 if page_size < 10 else 1000 if page_size > 1000 else page_size

    qry = SupportCase.query.order_by(SupportCase.id.desc())
    if getattr(current_user, "role", None) != "admin":
        qry = qry.filter(SupportCase.created_by == current_user.id)
    rows = qry.all()

    data = []
    for r in rows:
        row = {
            "الاسم": r.name or "",
            "الكود": r.code or "",
            "العمل المحقق": r.work_type or "",
            "أعمال دعم عامة": r.work_type != "حسابات بنكية" and (r.work_detail or "") or "",
            "اسم البريد المرسل": r.sender_email_name or "",
            "ملاحظات": r.notes or "",
            "رسالة التذكير": r.reminder_message or "",
            # 💥 يتم العرض بالتوقيت المحلي (التاريخ والوقت يظهران الآن)
            "وقت التذكير": _to_local_display(r.reminder_at),
            "Request Number": r.bank_request_number or "",
            "Bakery_Code": r.bank_bakery_code or "",
            "BANK_ID": r.bank_id or "",
            "BANK_ACC_NUMBER": r.bank_acc_number or "",
            "BANK_ACC_NAME": r.bank_acc_name or "",
            "National ID": r.bank_national_id or "",
            # 💥 يتم العرض بالتوقيت المحلي (التاريخ والوقت يظهران الآن)
            "تاريخ التسجيل": _to_local_display(r.created_at),
            "أنشأه": (r.creator.username if r.creator else ""),
            "ID": r.id
        }
        data.append(row)

    df = _df_text(pd.DataFrame(data))
    if df.empty:
        return render_template("support/index.html",
            title="الدعم الفني",
            is_admin=(getattr(current_user, "role", None) == "admin"),
            cols=[], rows=[], search_cols=[],
            q=q, search_in=search_in, page=page, page_size=page_size,
            pagination={"page": 1, "total_pages": 1},
            due_times=[]
        )

    # 💥 تأكيد الأعمدة الأساسية وترتيبها (لضمان ظهورها)
    cols_order = [
        "الاسم","الكود","العمل المحقق","أعمال دعم عامة",
        "اسم البريد المرسل","ملاحظات","رسالة التذكير",
        "وقت التذكير", # 💥 هذا هو العمود الذي نريده
        "Request Number","Bakery_Code","BANK_ID","BANK_ACC_NUMBER","BANK_ACC_NAME","National ID",
        "تاريخ التسجيل", # 💥 هذا هو العمود الآخر الذي نريده
        "أنشأه","ID"
    ]
    # التصفية على الأعمدة الموجودة فقط بالترتيب المطلوب
    final_cols = [c for c in cols_order if c in df.columns]
    
    # 💥 التعديل الأهم: إزالة الأعمدة الفارغة فقط للمسح البصري ولكن إبقاء الأعمدة الرئيسية 
    filtered = _filter_dataframe(df, q, search_in)
    
    # قائمة الأعمدة المعروضة للمستخدم (مع إخفاء ID)
    display_cols = [c for c in final_cols if c != "ID"]
    
    # تطبيق التصفية على الإطار النهائي
    visible_df = filtered[display_cols + ["ID"]].copy() 

    page_df, total = _paginate(visible_df, page, page_size)
    total_pages = max(1, (total + page_size - 1)//page_size)

    cols = [c for c in page_df.columns if c != "ID"]
    rows = page_df.to_dict(orient="records")

    def _page_url(n):
        return url_for("support_bp.index", q=q, search_in=search_in, page=n, page_size=page_size)

    first_pages = [n for n in [1,2,3] if n <= total_pages]
    pagination = {
        "prev": _page_url(page-1) if page>1 else None,
        "next": _page_url(page+1) if page<total_pages else None,
        "first_pages": [{"n":n,"url":_page_url(n),"active":(n==page)} for n in first_pages],
        "show_ellipsis": total_pages > 3,
        "last": {"n":total_pages,"url":_page_url(total_pages),"active":(page==total_pages)} if total_pages>3 else None,
        "page": page, "total_pages": total_pages
    }

    due_times = [r.get("وقت التذكير", "") for r in rows if r.get("وقت التذكير", "")]
    return render_template("support/index.html",
        title="الدعم الفني",
        is_admin=(getattr(current_user, "role", None) == "admin"),
        # 💥 إرسال قائمة الأعمدة النهائية للمقارنة في index.html
        cols=cols, 
        rows=rows, 
        search_cols=[c for c in visible_df.columns if c != "ID"],
        q=q, search_in=search_in, page=page, page_size=page_size,
        pagination=pagination, due_times=due_times
    )

@support_bp.route("/export")
@login_required
def export():
    q = request.args.get("q", "")
    search_in = request.args.get("search_in", "all")

    qry = SupportCase.query.order_by(SupportCase.id.desc())
    if getattr(current_user, "role", None) != "admin":
        qry = qry.filter(SupportCase.created_by == current_user.id)
    rows = qry.all()

    data = []
    for r in rows:
        # 💥 هنا نقوم بحفظ التواريخ كـ datetime object لكي يتم تصديرها بشكل صحيح في Excel
        reminder_dt = _parse_utc_str(r.reminder_at)
        created_dt = _parse_utc_str(r.created_at)
        
        row = {
            "الاسم": r.name or "",
            "الكود": r.code or "",
            "العمل المحقق": r.work_type or "",
            "أعمال دعم عامة": r.work_type!="حسابات بنكية" and (r.work_detail or "") or "",
            "اسم البريد المرسل": r.sender_email_name or "",
            "ملاحظات": r.notes or "",
            "رسالة التذكير": r.reminder_message or "",
            # 💥 التعديل: إظهار التاريخ والوقت المحلي كـ datetime object لـ Excel
            "وقت التذكير": (reminder_dt + timedelta(minutes=_TZ_OFFSET_MIN)) if reminder_dt else None,
            "Request Number": r.bank_request_number or "",
            "Bakery_Code": r.bank_bakery_code or "",
            "BANK_ID": r.bank_id or "",
            "BANK_ACC_NUMBER": r.bank_acc_number or "",
            "BANK_ACC_NAME": r.bank_acc_name or "",
            "National ID": r.bank_national_id or "",
            # 💥 التعديل: إظهار التاريخ والوقت المحلي كـ datetime object لـ Excel
            "تاريخ التسجيل": (created_dt + timedelta(minutes=_TZ_OFFSET_MIN)) if created_dt else None,
            "أنشأه": (r.creator.username if r.creator else "")
        }
        data.append(row)
        
    df = pd.DataFrame(data)
    
    # 💥 نقوم هنا بتحويل الأعمدة النصية فقط باستخدام _df_text
    # نحتفظ بأعمدة التاريخ كـ datetime object
    cols_to_textify = [c for c in df.columns if c not in ["وقت التذكير", "تاريخ التسجيل"]]
    # Apply _textify to the columns to be treated as text (this will remove decimal parts, etc.)
    for c in cols_to_textify:
        df[c] = df[c].map(_textify)

    # 💥 تأكيد نوع الأعمدة الزمنية كـ datetime لضمان تطبيق تنسيق Excel بشكل صحيح
    if "وقت التذكير" in df.columns:
        df["وقت التذكير"] = pd.to_datetime(df["وقت التذكير"], errors="coerce")
    if "تاريخ التسجيل" in df.columns:
        df["تاريخ التسجيل"] = pd.to_datetime(df["تاريخ التسجيل"], errors="coerce")

    # 💥 التأكيد على الأعمدة قبل التصدير
    base_cols = [
        "الاسم","الكود","العمل المحقق","أعمال دعم عامة",
        "اسم البريد المرسل","ملاحظات","رسالة التذكير","وقت التذكير",
        "Request Number","Bakery_Code","BANK_ID","BANK_ACC_NUMBER","BANK_ACC_NAME","National ID",
        "تاريخ التسجيل",
        "أنشأه"
    ]
    cols_order = [c for c in base_cols if c in df.columns]
    
    out = _drop_empty_columns(_filter_dataframe(df[cols_order], q, search_in))
    
    if out.empty:
        flash("لا توجد بيانات لتصديرها.", "warning")
        return redirect(url_for("support_bp.index", q=q, search_in=search_in))
    return _excel_response(out, "الدعم_الفني.xlsx")


# صلاحيات
def _ensure_owner_or_admin(rec: SupportCase):
    if getattr(current_user, "role", None) == "admin":
        return True
    return rec.created_by == current_user.id

# 💥💥 مسار API جديد للتحقق من التكرار 💥💥
# 💥 التعديل: تم تبسيط المنطق للتركيز على BANK_ACC_NUMBER وإرجاع رسالة للعميل
@support_bp.route("/check_bank_data", methods=["POST"])
@login_required
def check_bank_data():
    acc_num = (request.form.get("bank_acc_number") or "").strip()
    rid = request.form.get("record_id", type=int) # لمعرفة إذا كان تعديلاً

    # 1. التحقق من تكرار BANK_ACC_NUMBER (فقط)
    if acc_num:
        qry = SupportCase.query.filter(SupportCase.bank_acc_number == acc_num)
        
        # 💥💥 الأهم: استثناء السجل الحالي من التحقق 💥💥
        if rid:
            qry = qry.filter(SupportCase.id != rid)
            
        existing_acc = qry.first()
        if existing_acc:
            message = f"رقم الحساب البنكي **{acc_num}** مُسجَّل مسبقاً في سجل رقم {existing_acc.id}."
            # إرجاع حالة التكرار لتتم معالجتها بواسطة العميل (النافذة المنبثقة)
            return jsonify({"ok": False, "is_duplicate": True, "field": "bank_acc_number", "message": message, "existing_id": existing_acc.id})

    # 💥 إلغاء التحقق من تكرار Request Number من هذه الدالة، وتم التركيز على BANK_ACC_NUMBER حسب الطلب.
    return jsonify({"ok": True, "is_duplicate": False, "message": "البيانات غير مكررة."})

# إنشاء
@support_bp.route("/create", methods=["GET","POST"])
@login_required
def create():
    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        code = (request.form.get("code") or "").strip()
        work_type = (request.form.get("work_type") or "").strip()
        work_detail = (request.form.get("work_detail") or "").strip()
        bank_request_number = (request.form.get("bank_request_number") or "").strip()
        bank_bakery_code    = (request.form.get("bank_bakery_code") or "").strip()
        bank_id             = (request.form.get("bank_id") or "").strip()
        bank_acc_number     = (request.form.get("bank_acc_number") or "").strip()
        bank_acc_name       = (request.form.get("bank_acc_name") or "").strip()
        bank_national_id    = (request.form.get("bank_national_id") or "").strip()
        sender = (request.form.get("sender_email_name") or "").strip()
        notes = (request.form.get("notes") or "").strip()
        reminder_message = (request.form.get("reminder_message") or "").strip()
        reminder_at_local = (request.form.get("reminder_at_local") or "").strip() # 💥 الاسم الصحيح
        
        # 💥 إضافة حقل الإجبار على الحفظ 💥
        force_bank_save = request.form.get("force_bank_save", "0") == "1"

        if not name:
            flash("الاسم حقل إلزامي.", "warning"); return redirect(url_for("support_bp.create"))
        if code and not re.fullmatch(r"[\d\s\-/]+", code):
            flash("حقل الكود يجب أن يحتوي على أرقام أو مسافات أو شرطة (-) أو شرطة مائلة (/) فقط.", "warning")
            return redirect(url_for("support_bp.create"))
        if work_type not in ("أعمال دعم عامة","حسابات بنكية"):
            flash("اختر نوع عمل محقق صحيح.", "warning"); return redirect(url_for("support_bp.create"))
        if work_type == "حسابات بنكية" and not ((bank_acc_number or "").upper().startswith("EG")):
            pass 

        if work_type == "حسابات بنكية": work_detail = "بيانات بنكية مُعبأة"
        
        # 💥💥 منطق التحقق من التكرار (لـ BANK_ACC_NUMBER فقط) 💥💥
        # **هنا يجب أن يكون التحقق النهائي على مستوى الخادم**
        # 💥 التعديل: يتم تجاوز التحقق إذا تم تأكيد الحفظ مسبقاً من قبل العميل
        if work_type == "حسابات بنكية" and not force_bank_save:
            # 1. التحقق من تكرار BANK_ACC_NUMBER
            if bank_acc_number:
                existing_acc = SupportCase.query.filter(SupportCase.bank_acc_number == bank_acc_number).first()
                if existing_acc:
                    # إذا وصل هنا دون تأكيد، يعني أن العميل لم يطلب التأكيد، نعيد التوجيه لضمان سلامة البيانات
                    flash(f"رقم الحساب البنكي **{bank_acc_number}** مُسجَّل مسبقاً في سجل رقم {existing_acc.id}. يرجى تأكيد الحفظ مرة أخرى.", "danger")
                    return redirect(url_for("support_bp.create"))

            # 2. التحقق من تكرار Request Number - تم الإلغاء
            # تم حذف المنطق الخاص بالتحقق من تكرار bank_request_number
        # 💥 نهاية منطق التحقق من التكرار 💥

        reminder_at_utc = _to_utc_str_from_local(reminder_at_local) if reminder_at_local else ""
        # اضبط وقت التشغيل القادم فقط إذا كان في المستقبل بالنسبة لـ UTC
        next_fire_at = ""
        try:
            ra = _parse_utc_str(reminder_at_utc)
            if ra and ra > _utc_now():
                next_fire_at = reminder_at_utc
        except Exception:
            next_fire_at = ""

        rec = SupportCase(
            name=name, code=code, work_type=work_type, work_detail=work_detail,
            sender_email_name=sender, notes=notes,
            reminder_message=reminder_message, 
            reminder_at=reminder_at_utc,
            next_fire_at=next_fire_at, 
            dismissed=False, created_by=current_user.id,
            bank_request_number=bank_request_number, bank_bakery_code=bank_bakery_code,
            bank_id=bank_id, bank_acc_number=bank_acc_number,
            bank_acc_name=bank_acc_name, bank_national_id=bank_national_id
        )
        db.session.add(rec); db.session.commit()
        flash("تمت الإضافة.", "success")
        return redirect(url_for("support_bp.index"))

    return render_template("support/form.html", title="إضافة - الدعم الفني", mode="create", rec=None)

# تعديل
@support_bp.route("/<int:rid>/edit", methods=["GET","POST"])
@login_required
def edit(rid):
    rec = SupportCase.query.get_or_404(rid)
    if not _ensure_owner_or_admin(rec):
        flash("غير مسموح بتعديل هذا السجل.", "warning")
        return redirect(url_for("support_bp.index"))

    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        code = (request.form.get("code") or "").strip()
        work_type = (request.form.get("work_type") or "").strip()
        work_detail = (request.form.get("work_detail") or "").strip()

        bank_request_number = (request.form.get("bank_request_number") or "").strip()
        bank_bakery_code    = (request.form.get("bank_bakery_code") or "").strip()
        bank_id             = (request.form.get("bank_id") or "").strip()
        bank_acc_number     = (request.form.get("bank_acc_number") or "").strip()
        bank_acc_name       = (request.form.get("bank_acc_name") or "").strip()
        bank_national_id    = (request.form.get("bank_national_id") or "").strip()

        sender = (request.form.get("sender_email_name") or "").strip()
        notes = (request.form.get("notes") or "").strip()
        reminder_message = (request.form.get("reminder_message") or "").strip()
        
        reminder_at_local = (request.form.get("reminder_at_local") or "").strip() # 💥 الاسم الصحيح
        
        # 💥 إضافة حقل الإجبار على الحفظ 💥
        force_bank_save = request.form.get("force_bank_save", "0") == "1"


        if not name:
            flash("الاسم حقل إلزامي.", "warning"); return redirect(url_for("support_bp.edit", rid=rid))
        
        if code and not re.fullmatch(r"[\d\s\-/]+", code):
            flash("حقل الكود يجب أن يحتوي على أرقام أو مسافات أو شرطة (-) أو شرطة مائلة (/) فقط.", "warning")
            return redirect(url_for("support_bp.edit", rid=rid))
            
        if work_type not in ("أعمال دعم عامة","حسابات بنكية"):
            flash("اختر نوع عمل محقق صحيح.", "warning"); return redirect(url_for("support_bp.edit", rid=rid))

        if work_type == "حسابات بنكية": work_detail = "بيانات بنكية مُعبأة"
        
        # 💥💥 منطق التحقق من التكرار في التعديل (لـ BANK_ACC_NUMBER فقط) 💥💥
        # **هنا يجب أن يكون التحقق النهائي على مستوى الخادم**
        # 💥 التعديل: يتم تجاوز التحقق إذا تم تأكيد الحفظ مسبقاً من قبل العميل
        if work_type == "حسابات بنكية" and not force_bank_save:
            # 1. التحقق من تكرار BANK_ACC_NUMBER
            if bank_acc_number:
                existing_acc = SupportCase.query.filter(
                    SupportCase.bank_acc_number == bank_acc_number,
                    SupportCase.id != rid # استثناء السجل الحالي
                ).first()
                if existing_acc:
                    # إذا وصل هنا دون تأكيد، يعني أن العميل لم يطلب التأكيد، نعيد التوجيه لضمان سلامة البيانات
                    flash(f"رقم الحساب البنكي **{bank_acc_number}** مُسجَّل مسبقاً في سجل رقم {existing_acc.id}. يرجى تأكيد الحفظ مرة أخرى.", "danger")
                    return redirect(url_for("support_bp.edit", rid=rid))

            # 2. التحقق من تكرار Request Number - تم الإلغاء
            # تم حذف المنطق الخاص بالتحقق من تكرار bank_request_number
        # 💥 نهاية منطق التحقق من التكرار 💥

        rec.name = name
        rec.code = code
        rec.work_type = work_type
        rec.work_detail = work_detail
        rec.sender_email_name = sender
        rec.notes = notes
        rec.reminder_message = reminder_message

        reminder_at_utc = _to_utc_str_from_local(reminder_at_local) if reminder_at_local else ""
        
        rec.reminder_at = reminder_at_utc
        # اضبط وقت التشغيل القادم فقط إذا كان في المستقبل بالنسبة لـ UTC
        try:
            ra = _parse_utc_str(reminder_at_utc)
            rec.next_fire_at = (reminder_at_utc if (ra and ra > _utc_now()) else "")
        except Exception:
            rec.next_fire_at = ""
        rec.dismissed = False

        rec.bank_request_number = bank_request_number
        rec.bank_bakery_code    = bank_bakery_code
        rec.bank_id             = bank_id
        rec.bank_acc_number     = bank_acc_number
        rec.bank_acc_name       = bank_acc_name
        rec.bank_national_id    = bank_national_id

        db.session.commit()
        flash("تم الحفظ.", "success")
        return redirect(url_for("support_bp.index"))
    
    rec.reminder_at_local = ""
    utc_dt = _parse_utc_str(rec.reminder_at)
    if utc_dt:
        local_dt = utc_dt + timedelta(minutes=_TZ_OFFSET_MIN)
        rec.reminder_at_local = local_dt.strftime("%Y-%m-%dT%H:%M") 

    if rec.work_type == "حسابات بنكية":
        rec.category = "قسم الدعم الفني"
        rec.work_type = "حسابات بنكية"
        rec.type_select = "" 
    elif rec.work_type == "أعمال دعم عامة" and not rec.bank_request_number:
        rec.category = "قسم الدعم الفني"
        rec.work_type = "أعمال دعم عامة"
        rec.type_select = ""
    else:
        rec.category = ""
        rec.work_type = ""
        rec.type_select = ""

    return render_template("support/form.html", title="تعديل - الدعم الفني", mode="edit", rec=rec)

# حذف
@support_bp.route("/<int:rid>/delete", methods=["POST"])
@login_required
def delete(rid):
    rec = SupportCase.query.get_or_404(rid)
    if not _ensure_owner_or_admin(rec):
        flash("غير مسموح بحذف هذا السجل.", "warning")
        return redirect(url_for("support_bp.index"))
    db.session.delete(rec); db.session.commit()
    flash("تم الحذف.", "success")
    return redirect(url_for("support_bp.index"))

# API للتذكير
@support_bp.route("/reminders/poll")
@login_required
def reminders_poll():
    qry = SupportCase.query.filter_by(dismissed=False).filter(SupportCase.created_by == current_user.id)
    # قارن عند بداية الدقيقة لضمان الانطلاق الدقيق
    now_utc_min = _utc_now_floor_minute()
    item = None
    for r in qry.order_by(SupportCase.id.asc()).all():
        nxt = _parse_utc_str(r.next_fire_at) or _parse_utc_str(r.reminder_at)
        # انطلق فقط عندما يحين بداية الدقيقة المحددة أو بعدها
        if nxt and nxt <= now_utc_min:
            item = r
            break
    if not item:
        return jsonify({"ok": True, "reminder": None})
        
    return jsonify({
        "ok": True,
        "reminder": {
            "id": item.id,
            "message": item.reminder_message or "(بدون رسالة)",
            "at": item.next_fire_at or item.reminder_at or "",
            "name": item.name,
            "code": item.code,
            "work_type": item.work_type,
            "created_at": _to_local_display(item.created_at),
        }
    })

@support_bp.route("/reminders/snooze", methods=["POST"])
@login_required
def reminders_snooze():
    rid = request.form.get("id", type=int)
    mins = request.form.get("mins", default=30, type=int)
    r = SupportCase.query.get_or_404(rid)
    if r.created_by != current_user.id:
        return jsonify({"ok": False, "error": "forbidden"}), 403
    base = _parse_utc_str(r.next_fire_at) or _parse_utc_str(r.reminder_at) or _utc_now()
    r.next_fire_at = (base + timedelta(minutes=max(1, mins))).strftime("%Y-%m-%d %H:%M")
    r.dismissed = False
    db.session.commit()
    return jsonify({"ok": True})

@support_bp.route("/reminders/dismiss", methods=["POST"])
@login_required
def reminders_dismiss():
    rid = request.form.get("id", type=int)
    r = SupportCase.query.get_or_404(rid)
    if r.created_by != current_user.id:
        return jsonify({"ok": False, "error": "forbidden"}), 403
    r.dismissed = True
    r.next_fire_at = ""
    db.session.commit()
    return jsonify({"ok": True})

# دوال توافق خلفي (يمكن إبقاؤها)
@support_bp.get("/reminders/due")
@login_required
def reminders_due():
    # قارن عند بداية الدقيقة لضمان الانطلاق الدقيق
    now_utc_min = _utc_now_floor_minute()
    qry = SupportCase.query.filter_by(dismissed=False).filter(SupportCase.created_by == current_user.id)
    items = []
    for r in qry.all():
        nxt = _parse_utc_str(r.next_fire_at) or _parse_utc_str(r.reminder_at)
        if nxt and nxt <= now_utc_min:
            items.append({
                "id": r.id,
                "name": r.name or "",
                "code": r.code or "",
                "message": r.reminder_message or "",
                "when": r.reminder_at or ""
            })
    return jsonify({"items": items})

@support_bp.post("/reminders/<int:rid>/snooze")
@login_required
def reminders_snooze_id(rid):
    r = SupportCase.query.get_or_404(rid)
    if r.created_by != current_user.id:
        return jsonify({"ok": False, "error": "forbidden"}), 403
    base = _parse_utc_str(r.next_fire_at) or _parse_utc_str(r.reminder_at) or _utc_now()
    r.next_fire_at = (base + timedelta(minutes=30)).strftime("%Y-%m-%d %H:%M")
    r.dismissed = False
    db.session.commit()
    return jsonify({"ok": True})

@support_bp.post("/reminders/<int:rid>/dismiss")
@login_required
def reminders_dismiss_id(rid):
    r = SupportCase.query.get_or_404(rid)
    if r.created_by != current_user.id:
        return jsonify({"ok": False, "error": "forbidden"}), 403
    r.dismissed = True
    r.next_fire_at = ""
    db.session.commit()
    return jsonify({"ok": True})