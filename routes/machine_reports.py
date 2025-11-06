from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, current_app, jsonify
from flask_login import login_required, current_user
from models import db
from models_reports import ReportState, ServiceTicket
from utils.decorators import role_required, permission_required
import pandas as pd
import json
import io
import re
from decimal import Decimal
import numpy as np 
from io import BytesIO 
from datetime import datetime 
from time import time

machine_reports_bp = Blueprint('machine_reports_bp', __name__)

# أقسام التقارير العامة
CATEGORIES = {
    "bakeries": "مخابز",
    "ration": "تموين",
    "substitute": "الاستبدال",
}

# قائمة الأعطال المسموح بها (تظهر في القائمة المنسدلة في الواجهة)
ALLOWED_FAULT_TYPES = [
    "ريدر", "سوفت", "طباعه", "شحن", "سوكت", "شبكه", "شاشه", "بيت شريحه", "F2", "KEYS", "POWER"
]

# 💡 تم التعديل: نضمن أن مفاتيح "العميل" هي الأولوية القصوى للدمج (المفاتيح 2 و 4)
ENTITY_KEYS = {
    "bakeries": [("رقم العميل", "اسم العميل"), ("رقم المخبز", "اسم المخبز")],
    "default":  [("رقم العميل", "اسم العميل"), 
                 ("رقم التاجر", "اسم التاجر"),
                 ("trader_id", "trader_name")]
}

# مفاتيح الملف الثالث
OFFICE_KEYS = ("الادارة", "المكتب")
# السماح بستة ملفات في الاستيراد بحسب المتطلبات الجديدة
MAX_FILES = 6

# كاش خفيف لنتائج تحميل سجل الزيارات لتقليل ضربات قاعدة البيانات خلال فترة قصيرة
_VISIT_CACHE = {'df': None, 'meta': None, 'ts': 0}
_VISIT_CACHE_TTL_SEC = 60

# الأعمدة الخاصة بتفاصيل مسلسل الماكينة (ستظهر تحت بعضها في الشاشة)
MACHINE_DETAIL_COLS = [
    'مسلسل الماكينة',
    'ماكينة رئيسية/فرعية',
    'رقم الماكينة',
    'حالة الماكينة',
    'شريحة 1',  
    'شريحة 2',  
    'حالة نظام المطحن',
    'SW_AC_SUP',
    'SW_IC_SUP',
    'SW_OD_SUP',
    'POS_VERSION',
    'اسم الخبز',   
    'LOAF_BALANCE1',
    'ساعة بدء البيع',
    'ساعة نهاية البيع',
]

# 💡 الأعمدة التي يتوقعها HTML في المربع الأول 'البيانات الأساسية'
CUSTOMER_DETAIL_COLS = [
     'الادارة', 'المكتب', 'رقم العميل', 'اسم العميل', 'اسم المسؤل',
     'الرقم القومي', 'رقم المحمول', 'رقم هاتف', 'حالة العميل', 'ملاحظات'
]

# ==================== (1) الدوال المساعدة العامة (Utils, Coercion, State) ====================

# ========== تحويل كل القيم لنص + إزالة الوقت + تنظيف nan/none ==========
_EMPTY_TOKENS = {"nan", "none", "null", "na", "n/a", "nat", "-", "—"}
_DATE_RE_1 = re.compile(r"^(\d{4}[-/]\d{1,2}[-/]\d{1,2})[ T]\d{1,2}:\d{2}(?::\d{2})?$")
_DATE_RE_2 = re.compile(r"^(\d{1,2}[-/]\d{1,2}[-/]\d{4})[ T]\d{1,2}:\d{2}(?::\d{2})?$")

def _strip_time_from_date(txt: str) -> str:
    m = _DATE_RE_1.match(txt)
    if m: return m.group(1)
    m = _DATE_RE_2.match(txt)
    if m: return m.group(1)
    return txt

def _textify(v) -> str:
    """تحويل القيمة إلى نص نظيف مع معالجة الأرقام العشرية والأسية بتبسيط."""
    if v is None: return ""
    
    if isinstance(v, (float, np.float64)):
        if np.isinf(v):
            return "" 
        if np.isnan(v):
            return "" 
    
    t = str(v).strip()
    if t == "" or t.lower() in _EMPTY_TOKENS: return ""
    
    t = t.replace(",", "")
    # لا نقوم بإزالة وقت التاريخ بعد الآن لضمان ظهور وقت الزيارة في الواجهة

    # 💡 FIX: تجنب تحويل أرقام التعريف الطويلة (مثل المسلسلات أو الشرائح) إلى float لتجنب فقدان الدقة
    # إذا كان النص يتكون من أرقام فقط وطوله أكبر من 12، فاحفظه كنص (String).
    if t.isdigit() and len(t) > 12: 
        return t 
        
    try:
        float_val = float(t)
        
        if np.isinf(float_val) or np.isnan(float_val):
            return ""
            
        if float_val == int(float_val):
            return str(int(float_val))
        
        # محاولة التعامل مع الأرقام العشرية الكبيرة بطريقة Decimal
        try:
             # تحويلها إلى قيمة صحيحة مقربة (إن أمكن) لتجنب الأصفار العشرية غير الضرورية
             # أو تركها كقيمة عشرية مبسطة إن لم يكن رقماً صحيحاً
             return str(float_val) 
        except Exception:
             pass

    except ValueError:
        pass 

    return t


def _coerce_text_df(df: pd.DataFrame) -> pd.DataFrame:
    # 💡 تم التعديل هنا: إضافة تحقق صريح باستخدام isinstance لحل مشكلة 'function' object has no attribute 'empty'
    if df is None or not isinstance(df, pd.DataFrame) or df.empty: 
        return pd.DataFrame()
        
    out = df.copy()
    
    # 1. تنظيف أسماء الأعمدة
    out.columns = [str(c).strip().replace("\n"," ").replace("\r"," ") for c in out.columns]
    
    # 2. ضمان تفرد أسماء الأعمدة (قد تكون الأعمدة المكررة تسبب مشاكل)
    seen = set()
    new_cols = []
    for col in out.columns:
        if col in seen:
            continue
        seen.add(col)
        new_cols.append(col)
        
    out = out[new_cols] 
    
    # 3. تطبيق _textify
    for c in out.columns:
        out[c] = out[c].map(_textify)
        
    return out

# ========== State (الحالة) ==========
def _ensure_tables():
    """وظيفة تأكد من وجود الجداول (لـ ReportState)"""
    # 💡 ملاحظة: لا يجب استدعاء db.create_all() هنا بل في نقطة تهيئة التطبيق
    # ولكن تم تركها مؤقتاً لتجنب فشل إذا لم يكن هناك تهيئة عامة.
    try: db.create_all() 
    except Exception: pass

def _df_to_json(df: pd.DataFrame) -> str:
    return _coerce_text_df(df).to_json(orient="records", force_ascii=False)


def _json_to_df(js: str) -> pd.DataFrame:
    if not js: return pd.DataFrame()
    df = pd.DataFrame(json.loads(js))
    return _coerce_text_df(df)

def _load_state(category: str):
    """
    تحميل سجل حالة التقرير للفئة المحددة.
    - أولوية: حالة مرتبطة بالمستخدم الحالي.
    - سقوط احتياطي: حالة عامة بدون user_id إذا لم توجد حالة المستخدم.
    """
    _ensure_tables()
    try:
        # أولًا: حالة المستخدم إن كان مسجّلًا
        if current_user.is_authenticated:
            row = ReportState.query.filter_by(category=category, user_id=current_user.id).first()
            if row:
                return row
        # ثانيًا: سقوط احتياطي على حالة عامة (قد تكون محفوظة من خدمات التجار بدون user_id)
        # نأخذ أحدث سجل بنفس الفئة بغض النظر عن user_id
        return (ReportState.query
                .filter(ReportState.category == category)
                .order_by(ReportState.id.desc())
                .first())
    except Exception:
        return None

def _save_state(category: str, df: pd.DataFrame = None, mapping: dict = None):
    """
    حفظ سجل حالة التقرير الخاص بالمستخدم الحالي والفئة المحددة.
    """
    if not current_user.is_authenticated:
        return
        
    # 💡 تم التعديل: يستخدم user_id
    row = ReportState.query.filter_by(category=category, user_id=current_user.id).first()
    if not row:
        row = ReportState(category=category, user_id=current_user.id)
        db.session.add(row)
        
    if df is not None:
        row.data_json = _df_to_json(df)
    if mapping is not None:
        row.mapping_json = json.dumps(mapping, ensure_ascii=False)
        
    db.session.commit()

def _apply_mapping(df: pd.DataFrame, mapping: dict) -> pd.DataFrame:
    """تطبيق إعادة تسمية وترتيب الأعمدة"""
    if df is None or df.empty or not mapping:
        return df
    rename = mapping.get("rename") or {}
    order  = [c for c in (mapping.get("order") or []) if c]
    out = df.copy()
    if rename: out = out.rename(columns=rename)
    if order:
        front = [c for c in order if c in out.columns]
        # إظهار الأعمدة المحفوظة فقط وإخفاء الباقي من العرض/التصدير
        out = out[front] if front else out
    return _coerce_text_df(out)

def _drop_empty_columns(df: pd.DataFrame) -> pd.DataFrame:
    """إزالة الأعمدة التي لا تحتوي على أي بيانات"""
    if df is None or df.empty: return df
    
    def _is_empty_series(s: pd.Series) -> bool:
        if not isinstance(s, pd.Series):
             return s.empty if hasattr(s, 'empty') else True
             
        vals = s.fillna("").astype(str).str.strip().str.lower()
        return ((vals == "") | (vals.isin(_EMPTY_TOKENS))).all()
    
    keep = [c for c in df.columns if not _is_empty_series(df[c])]
    return df[keep] if keep else df.iloc[:, 0:0]

def _filter_dataframe(df: pd.DataFrame, query: str, search_cols: list[str] | None = None) -> pd.DataFrame:
    """تصفية سريعة مع توحيد عربي، بأقل عدد من العمليات.
    - تجمع نص الصف مرة واحدة عبر الأعمدة المستهدفة ثم تطبق contains.
    - تقلل التكلفة من O(rows * cols) إلى O(rows).
    """
    if not query:
        return df

    q = _norm_key_text(query).lower()
    if not q:
        return df

    # الأعمدة المستهدفة للتصفية
    cols_to_search = (
        [c for c in (search_cols or []) if c in df.columns]
        if (search_cols and isinstance(search_cols, list)) else list(df.columns)
    )

    if not cols_to_search:
        return pd.DataFrame(columns=df.columns)

    # المسار السريع: تجميع نص الصف مرة واحدة ثم التوحيد والبحث
    try:
        # تحويل إلى نص وتعبئة الفراغات ثم تجميع الصفوف
        rows_text = df[cols_to_search].fillna("").astype(str)
        # تجميع نص الصف مرة واحدة
        all_text = rows_text.apply(lambda r: " ".join(r.values.tolist()), axis=1)
        # توحيد عربي + تصغير ثم contains
        all_text_norm = all_text.map(_norm_key_text).str.lower()
        mask = all_text_norm.str.contains(q, na=False)
        return df[mask]
    except Exception:
        # مسار احتياطي: نفس المنهج السابق عمودًا بعمود
        try:
            df_text_normalized = df[cols_to_search].astype(str).apply(
                lambda col: col.map(lambda v: _norm_key_text(v).lower())
            )
            mask = df_text_normalized.apply(lambda col: col.str.contains(q, na=False))
            return df[mask.any(axis=1)]
        except Exception:
            # في حال حدوث خطأ غير متوقع، أعد الإطار كما هو لتجنب كسر الواجهة
            return df


# ==================== (2) وظائف البحث والتجميع ====================

def _get_entity_grouping_keys(filtered_df: pd.DataFrame, category: str) -> list:
    """يحدد أعمدة رقم واسم الكيان (العميل/المخبز/التاجر) لاستخدامها كمفاتيح تجميع بشكل أكثر مرونة."""

    df_cols = list(filtered_df.columns)

    # أزواج مفاتيح محتملة بحسب الفئة مع تضمين المرادفات العربية والإنجليزية
    candidate_pairs: list[tuple[str, str]] = []
    if category == 'bakeries':
        candidate_pairs.extend([
            ('رقم العميل', 'اسم العميل'),
            ('رقم المخبز', 'اسم المخبز'),
            ('رقم التاجر', 'اسم التاجر'),
        ])
    elif category in ['ration', 'substitute']:
        candidate_pairs.extend([
            ('رقم العميل', 'اسم العميل'),
            ('رقم التاجر', 'اسم التاجر'),
        ])
    else:
        candidate_pairs.extend([
            ('رقم العميل', 'اسم العميل'),
            ('رقم التاجر', 'اسم التاجر'),
        ])

    # إضافة مرادفات عامة
    candidate_pairs.extend([
        ('Customer Code', 'Customer Name'),
        ('Customer ID', 'Customer Name'),
        ('Customer_ID', 'Customer Name'),
        ('Trader ID', 'Trader Name'),
        ('trader_id', 'trader_name'),
        ('رقم المخبز', 'اسم المخبز'),
    ])

    # اختر أول زوج موجود بالكامل في الأعمدة
    for code_key, name_key in candidate_pairs:
        if (code_key in df_cols) and (name_key in df_cols):
            return [code_key, name_key]

    # محاولة أخيرة: اكتشاف أي عمود يحتوي "رقم" وأي عمود يحتوي "اسم"
    code_key = next((c for c in df_cols if ('رقم' in c) or (c.lower() in {'customer code','customer id','customer_id','trader id','trader_id'})), None)
    name_key = next((c for c in df_cols if ('اسم' in c) or (c.lower() in {'customer name','trader name','trader_name'})), None)
    keys = []
    if code_key: keys.append(code_key)
    if name_key and name_key != code_key: keys.append(name_key)
    if keys:
        return keys

    # Fallback: أول عمودين إذا لم يتم العثور على مفاتيح مناسبة
    return list(df_cols)[:2]

# 💡 تم تحديث MACHINE_DATA_SOURCE_MAPPING لضمان شمول بدائل الشرائح
MACHINE_DATA_SOURCE_MAPPING = {
    'مسلسل الماكينة': ['مسلسل الماكينة', 'مسلسل', 'Serial'], 
    'رقم الماكينة': ['رقم الماكينة', 'كود الماكينة', 'Machine Code'],
    'ماكينة رئيسية/فرعية': ['ماكينة رئيسية/فرعية', 'النوع', 'Type'], 
    'حالة الماكينة': ['حالة الماكينة', 'الحالة', 'Status'],
    # إلغاء استخدام CS_1/CS_2 كبدائل لأنها ليست أرقام شرائح فعلية
    'شريحة 1': ['شريحة 1', 'شريحة1', 'SIM1'], 
    'شريحة 2': ['شريحة 2', 'شريحة2', 'SIM2'], 
}
MACHINE_CODE_COL = 'رقم الماكينة'
MACHINE_SERIAL_COL = 'مسلسل الماكينة' 
SLICE_COLS = ['شريحة 1', 'شريحة 2']

# ========== وظائف مساعدة لاختيار مفاتيح الدمج المرنة (FIX للشرائح) ==========

def _find_actual_col(standard_col_name: str, df_cols: list) -> str | None:
    """يحدد الاسم الفعلي للعمود بناءً على قائمة الأسماء البديلة (في MACHINE_DATA_SOURCE_MAPPING) والأعمدة الموجودة في DataFrame."""
    # استخدام الاسم القياسي نفسه وأي بدائل محددة له
    source_names = MACHINE_DATA_SOURCE_MAPPING.get(standard_col_name, [standard_col_name])
    
    # الأولوية: الاسم القياسي أولاً، ثم البدائل
    for name in source_names:
        if name in df_cols:
            return name
    return None

def _pick_machine_keys(df1_cols: list, df2_cols: list) -> list[str]:
    """يحدد مفاتيح الماكينة المشتركة (مسلسل/رقم) للدمج بالاعتماد على أسماء الأعمدة الفعلية."""
    
    # محاولة إيجاد اسم العمود الفعلي لـ 'مسلسل الماكينة' في كل من DF1 و DF2
    serial1 = _find_actual_col(MACHINE_SERIAL_COL, df1_cols)
    serial2 = _find_actual_col(MACHINE_SERIAL_COL, df2_cols)
    
    # محاولة إيجاد اسم العمود الفعلي لـ 'رقم الماكينة' في كل من DF1 و DF2
    code1 = _find_actual_col(MACHINE_CODE_COL, df1_cols)
    code2 = _find_actual_col(MACHINE_CODE_COL, df2_cols)

    # الأولوية: مسلسل الماكينة
    if serial1 and serial2 and _norm_key_text(serial1) == _norm_key_text(serial2):
         return [serial1]
    
    # الأولوية التالية: رقم الماكينة
    if code1 and code2 and _norm_key_text(code1) == _norm_key_text(code2):
         return [code1]
         
    return [] # لا يوجد مفتاح ماكينة مشترك يمكن استخدامه


def _group_search_results(filtered_df: pd.DataFrame, category: str) -> list[dict]:
    """تجميع سجلات DataFrame"""
    if filtered_df.empty: return []

    group_keys = _get_entity_grouping_keys(filtered_df, category)
    
    valid_group_keys = [k for k in group_keys if k in filtered_df.columns]
    
    if len(valid_group_keys) < 2:
        # إذا لم يتم العثور على مفاتيح كيان ثنائية (رقم واسم)، نعتبر كل صف كياناً مستقلاً
        grouped = [(i, filtered_df.iloc[[i]]) for i in range(len(filtered_df))]
        valid_group_keys = []
    else:
        grouped = filtered_df.groupby(valid_group_keys, dropna=False).__iter__()

    result_list = []
    
    EXCLUDE_FROM_COMMON = [
        # أعمدة المسلسلات
        *MACHINE_DETAIL_COLS, 
        # أعمدة يجب استبعادها من البيانات المشتركة وتظهر في مكان آخر أو غير ضرورية
        'CS_3', 'COUNT_DIST', 'LOAF_BALANCE', 
        'timestamp', 'report_data',
        'CS_1', 'CS_2', # يتم استبعادها لأنها أصبحت بدائل لـ شريحة 1 و شريحة 2
    ] 
    
    for _, group_df in grouped:
        first_record = group_df.iloc[0].to_dict()
        common_data = {}
        
        # استخراج البيانات المشتركة (تظهر مرة واحدة للكيان)
        for col in group_df.columns:
              if col in first_record and col not in EXCLUDE_FROM_COMMON:
                  common_data[col] = first_record[col]

        # 💡 FIX: منطق استخلاص الشرائح عبر كود الماكينة (رقم الماكينة)
        # 1. إنشاء جدول بحث للشرائح بناءً على رقم الماكينة في المجموعة بأكملها
        slice_lookup = {}
        # تحديد اسم عمود "رقم الماكينة" الفعلي في بيانات العميل (Group_df)
        actual_machine_code_col = next((c for c in MACHINE_DATA_SOURCE_MAPPING[MACHINE_CODE_COL] if c in group_df.columns), MACHINE_CODE_COL)

        for _, r in group_df.iterrows():
            machine_code = _textify(r.get(actual_machine_code_col))
            if machine_code:
                row_slices = {}
                for slice_col in SLICE_COLS:
                    value = r.get(slice_col) # 1. تحقق من الاسم الأساسي
                    
                    # 2. تحقق من أسماء الأعمدة البديلة في نفس الصف
                    if _textify(value) == '':
                        for alt_col in MACHINE_DATA_SOURCE_MAPPING.get(slice_col, []):
                            alt_value = r.get(alt_col)
                            if _textify(alt_value) != '':
                                value = alt_value
                                break
                                
                    if _textify(value) != '':
                         row_slices[slice_col] = value

                # دمج النتائج: الأولوية للقيمة غير الفارغة المكتشفة أولاً
                if machine_code not in slice_lookup:
                     slice_lookup[machine_code] = row_slices
                else:
                     for k, v in row_slices.items():
                          if _textify(v) != '':
                               slice_lookup[machine_code][k] = v


        # استخراج تفاصيل الماكينات
        machine_details = []
        for _, row in group_df.iterrows():
            detail = {}
            
            current_machine_code = _textify(row.get(actual_machine_code_col))
            
            for col in MACHINE_DETAIL_COLS:
                # 1. القيمة الأساسية
                value = row.get(col) or '-' 
                
                # 2. منطق الشرائح: البحث في جدول الـ Lookup إذا كانت القيمة مفقودة
                if col in SLICE_COLS:
                    # إذا كانت القيمة الأساسية فارغة، نبحث في جدول الـ Lookup المجمّع
                    if _textify(value) == '' and current_machine_code and current_machine_code in slice_lookup:
                        lookup_value = slice_lookup[current_machine_code].get(col)
                        if _textify(lookup_value) != '':
                             value = lookup_value
                             
                # 3. البحث في البدائل الأخرى (يشمل المسلسل ورقم الماكينة وغيرهما)
                if _textify(value) == '' and col in MACHINE_DATA_SOURCE_MAPPING:
                    for alt_col in MACHINE_DATA_SOURCE_MAPPING[col]:
                        alt_value = row.get(alt_col)
                        if _textify(alt_value) != '':
                            value = alt_value
                            break
                            
                # 4. حفظ القيمة النهائية
                detail[col] = _textify(value) or '-'
                
            machine_details.append(detail)
            
        result_list.append({
            'common_data': common_data,
            'machine_details': machine_details,
            'group_keys': valid_group_keys,
        })

    return result_list

# ==================== (3) الدوال المساعدة لجلب البيانات الخارجية - محاكاة مؤقتة ====================

def _fetch_visit_data(customer_code: str, visit_history_df: pd.DataFrame, visit_period: str = 'month', month_label: str | None = None, year_label: str | None = None) -> dict:
    """
    حساب عدد الزيارات للشهر الحالي والسنة الحالية بشكل منفصل اعتمادًا على عمود التاريخ الموحّد 'التاريخ'.
    - إذا توفر التاريخ: يتم العد بدقة لكل فترة، مع تفاصيل حسب 'مسلسل' إن وُجدت.
    - إذا لم يتوفر التاريخ: يتم إرجاع إجمالي واحد ويُستخدم لكلتا الفترتين (Fallback).
    """
    if visit_history_df is None or visit_history_df.empty:
        return {'current_month': {'total': 0, 'details': {}}, 'current_year': {'total': 0, 'details': {}}}

    df = visit_history_df.copy()
    now = datetime.now()

    # محاولة استخراج التاريخ إذا كان موجودًا
    has_date = 'التاريخ' in df.columns
    has_serial = 'مسلسل' in df.columns

    if has_date:
        # تحويل التاريخ إلى datetime بأمان
        try:
            dates = pd.to_datetime(df['التاريخ'], errors='coerce', dayfirst=True)
        except Exception:
            dates = pd.to_datetime(df['التاريخ'].astype(str), errors='coerce', dayfirst=True)
        df['_dt'] = dates
        # بناء أقنعة الفترة
        # دعم اختيار شهر/سنة محددين من تسميات الملفات (_الفترة) إذا كانت متوفرة
        if month_label and isinstance(month_label, str) and len(month_label) == 7:
            try:
                target_year = int(month_label.split('-')[0])
                target_month = int(month_label.split('-')[1])
            except Exception:
                target_year = now.year
                target_month = now.month
        else:
            target_year = now.year
            target_month = now.month

        if year_label and isinstance(year_label, str) and len(year_label) == 4 and year_label.isdigit():
            try:
                target_year_for_year = int(year_label)
            except Exception:
                target_year_for_year = now.year
        else:
            target_year_for_year = target_year

        month_mask = (df['_dt'].dt.year == target_year) & (df['_dt'].dt.month == target_month)
        year_mask  = (df['_dt'].dt.year == target_year_for_year)

        # اختيار مجموعات بحسب الفترة المطلوبة
        if visit_period == 'year':
            # عند اختيار السنة، لا نُهمل حساب الشهر: إذا توفرت تسمية شهر أو أمكن اشتقاقها من التاريخ
            if '_الفترة' in df.columns and month_label:
                df_month = df[df['_الفترة'] == month_label]
            else:
                df_month = df[(df['_dt'].dt.year == target_year) & (df['_dt'].dt.month == target_month)]
            df_year  = df[year_mask]
        elif visit_period == 'recent_program':
            # البيانات الحديثة: اعتبر كل الصفوف حديثة ضمن لوحة واحدة
            df_month = df
            df_year  = df.iloc[0:0]
        else:  # الافتراضي 'month'
            # إذا كانت لدينا أعمدة تسميات الفترات، نعتمد تسمية الملف مباشرة للشهر المحدد
            if '_الفترة' in df.columns and month_label:
                df_month = df[df['_الفترة'] == month_label]
            else:
                df_month = df[month_mask]
            df_year  = df[year_mask]

        month_total = int(df_month.shape[0])
        year_total  = int(df_year.shape[0])

        # حساب آخر وقت زيارة متاح داخل الفترة المختارة (الشهر/الحديث)
        latest_dt_str = ""
        latest_by_serial = {}
        latest_serial = ""
        try:
            valid_month = df_month[df_month['_dt'].notna()] if ('_dt' in df_month.columns) else df_month
            if not valid_month.empty and ('التاريخ' in valid_month.columns):
                idx_latest = valid_month['_dt'].idxmax() if ('_dt' in valid_month.columns) else valid_month.index[-1]
                # حفظ النص الأصلي للتاريخ حتى لو كان يحتوي على الوقت
                latest_dt_str = str(valid_month.at[idx_latest, 'التاريخ']).strip()
                # حفظ المسلسل المرتبط بآخر سجل (إن وُجد)
                try:
                    if has_serial and ('مسلسل' in valid_month.columns):
                        latest_serial = str(valid_month.at[idx_latest, 'مسلسل']).strip()
                except Exception:
                    pass
                # حساب آخر وقت لكل مسلسل إذا توفّر عمود المسلسل
                if has_serial:
                    try:
                        for serial, sub in valid_month.groupby('مسلسل'):
                            sub_valid = sub[sub['_dt'].notna()] if ('_dt' in sub.columns) else sub
                            if not sub_valid.empty:
                                idx_s = sub_valid['_dt'].idxmax() if ('_dt' in sub_valid.columns) else sub_valid.index[-1]
                                latest_by_serial[serial] = str(sub_valid.at[idx_s, 'التاريخ']).strip()
                    except Exception:
                        pass
        except Exception:
            latest_dt_str = latest_dt_str or ""

        if has_serial:
            month_details = df_month.groupby('مسلسل').size().to_dict()
            year_details  = df_year.groupby('مسلسل').size().to_dict()
        else:
            month_details = {}
            year_details  = {}

        return {
            'current_month': {'total': month_total, 'details': month_details},
            'current_year':  {'total': year_total,  'details': year_details},
            'latest_datetime': latest_dt_str,
            'latest_serial': latest_serial,
            'latest_serial_times': latest_by_serial
        }
    else:
        # لا يوجد تاريخ: إجمالي واحد يُستخدم لكلا الفترتين
        total_count = int(df.shape[0])
        if has_serial:
            serial_counts = df.groupby('مسلسل').size().to_dict()
        else:
            serial_counts = {}
        return {
            'current_month': {'total': total_count, 'details': serial_counts},
            'current_year':  {'total': total_count, 'details': serial_counts}
        }

# ==================== إعادة بناء منطق الزيارات: توحيد الأعمدة والمطابقة ====================
def _standardize_visit_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    توحيد أعمدة سجل الزيارات إلى أسماء قياسية:
    - التاريخ → 'التاريخ'
    - المسلسل → 'مسلسل'
    - رقم العميل → 'رقم العميل'
    - اسم العميل → 'اسم العميل'
    يدعم أسماء بديلة شائعة ويُبقي الأعمدة الأخرى كما هي.
    """
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    cols = list(out.columns)
    # توسيع المرشحات لتغطية بدائل أكثر و البحث بالاشتقاق
    date_cands = [
        'التاريخ', 'تاريخ الزيارة', 'Date', 'Visit Date', 'تاريخ',
        'visit_date', 'date', 'created_at', 'dt'
    ]
    # توسيع المرادفات لتغطية مزيد من الحالات القادمة من ملفات أو خرائط مختلفة
    serial_cands = [
        'مسلسل', 'مسلسل الماكينة', 'Serial', 'POS Serial', 'POS', 'رقم الماكينة',
        'serial', 'pos serial', 'sn', 'POS_SN'
    ]
    code_cands = [
        'رقم العميل', 'Customer Code', 'Customer_ID', 'client_code', 'trader_id',
        'code', 'ID', 'id', 'Bakery ID', 'رقم المخبز', 'رقم التاجر'
    ]
    name_cands = [
        'اسم العميل', 'اسم المخبز', 'اسم التاجر', 'Customer Name', 'Trader Name', 'Bakery Name', 'trader_name',
        'name', 'customer_name', 'trader_name_ar'
    ]
    # إضافة مرادفات لعمود النوع لضمان الفلترة الصحيحة حسب القسم
    type_cands = ['النوع', 'نوع', 'Type', 'type', 'Category', 'category']

    def pick(cands):
        for c in cands:
            if c in cols:
                return c
        return None

    ren = {}
    # محاولة التقاط عمود التاريخ بالاشتقاق إذا لم يُعثر عليه صراحةً
    def fuzzy_pick_date(cols_list: list[str]) -> str | None:
        lowered = [str(c).lower() for c in cols_list]
        for i, lc in enumerate(lowered):
            if ('تاريخ' in lc) or ('اريخ' in lc) or ('date' in lc) or ('visit' in lc and 'date' in lc):
                return cols_list[i]
        return None

    # إلتقاط مرن لعمود الرقم (code/id) إذا لم يُعثر عليه مباشرةً
    def fuzzy_pick_code(cols_list: list[str]) -> str | None:
        lowered = [str(c).lower() for c in cols_list]
        for i, lc in enumerate(lowered):
            if ('code' in lc) or (lc == 'id') or ('trader' in lc and 'id' in lc) or ('bakery' in lc and 'id' in lc) or ('رقم' in lc and 'قومي' not in lc and 'مسلسل' not in lc):
                return cols_list[i]
        return None

    # إلتقاط مرن لعمود المسلسل إذا كانت الأسماء بحروف صغيرة أو مختصرة
    def fuzzy_pick_serial(cols_list: list[str]) -> str | None:
        lowered = [str(c).lower() for c in cols_list]
        for i, lc in enumerate(lowered):
            if ('serial' in lc) or ('sn' in lc) or ('pos' in lc and 'serial' in lc) or ('مسلسل' in lc):
                return cols_list[i]
        return None

    d = pick(date_cands) or fuzzy_pick_date(cols)
    s = pick(serial_cands) or fuzzy_pick_serial(cols)
    c = pick(code_cands) or fuzzy_pick_code(cols)
    n = pick(name_cands)
    tcol = pick(type_cands)
    if d: ren[d] = 'التاريخ'
    if s: ren[s] = 'مسلسل'
    if c: ren[c] = 'رقم العميل'
    if n: ren[n] = 'اسم العميل'
    if tcol: ren[tcol] = 'النوع'
    if ren:
        out = out.rename(columns=ren)

    # إذا لم يتم التعرف على التاريخ، جرّب اكتشافه حسب النوع: إذا وُجد عمود datetime أو عمود يُمكن تحويل غالب قيمه لتاريخ
    if 'التاريخ' not in out.columns:
        try:
            # مرشح: عمود نوعه datetime مباشرةً
            for c in out.columns:
                if np.issubdtype(out[c].dtype, np.datetime64):
                    out = out.rename(columns={c: 'التاريخ'})
                    break
            # إن لم يُلتقط، جرّب أفضل عمود يمكن تحويله
            if 'التاريخ' not in out.columns:
                best_col = None; best_score = -1
                for c in out.columns:
                    ser = out[c]
                    # تقييم قابلية تحويل النصوص/الأرقام إلى تاريخ
                    try:
                        # ترجمة أرقام عربية وتقليم
                        tmp = ser.astype(str).str.translate(_AR_DIGITS).str.strip()
                        dt1 = pd.to_datetime(tmp, errors='coerce', dayfirst=True)
                        dt2 = pd.to_datetime(tmp, errors='coerce', dayfirst=False)
                        score = max(dt1.notna().sum(), dt2.notna().sum())
                    except Exception:
                        score = -1
                    # دعم التواريخ الرقمية على طريقة Excel
                    try:
                        # إذا كان العمود رقميًا في معظم صفوفه
                        num_ratio = pd.to_numeric(ser, errors='coerce').notna().mean()
                        if num_ratio > 0.6:
                            dt_num = pd.to_datetime(pd.to_numeric(ser, errors='coerce'), unit='d', origin='1899-12-30', errors='coerce')
                            score = max(score, dt_num.notna().sum())
                    except Exception:
                        pass
                    if score > best_score:
                        best_score = score; best_col = c
                if best_col and best_score > 0:
                    out = out.rename(columns={best_col: 'التاريخ'})
        except Exception:
            pass

    # تحويل الأنواع إلى نص لحقول المطابقة لتجنّب عدم التطابق بسبب الأنواع
    for k in ['رقم العميل', 'اسم العميل', 'مسلسل']:
        if k in out.columns:
            try:
                out[k] = out[k].astype(str).str.strip()
            except Exception:
                pass
    return _drop_empty_columns(out)


def _inquiry_search(category: str, search_type: str, query: str, visit_period: str = 'recent_program') -> dict:
    """تنفيذ البحث السريع داخل بيانات التقرير المخزنة وإعادة هيكلة النتائج.
    تم تحسين الأداء باستخدام فهارس كاش في الذاكرة لمسار سريع، مع مسار احتياطي للتصفية التقليدية عند الحاجة.
    """
    
    # استخدم الكاش المُفهرس بدلاً من إعادة بناء المابنج في كل طلب
    cached = _get_inquiry_cache(category)
    # تجنب تقييم الحقيقة الغامض لـ DataFrame عند استخدام "or"
    mapped_df = cached.get('df') if (cached.get('df') is not None) else pd.DataFrame()
    all_cols = cached.get('cols') or []
    indexes = cached.get('indexes') or {}

    if mapped_df.empty:
        return {'success': False, 'message': f'لا توجد بيانات مستوردة لقسم {CATEGORIES.get(category, category)}.', 'items': []}
    
    # 1. تحديد أعمدة البحث بناءً على النوع (Search_Type)
    target_cols = []
    
    if search_type == 'code':
        # الأولوية للبحث برقم العميل ثم المخبز ثم التاجر
        target_cols_priority = ['رقم العميل', 'رقم المخبز', 'رقم التاجر', 'رقم الماكينة']
        for col in target_cols_priority:
            if col in all_cols:
                target_cols.append(col)
        if not target_cols:
             target_cols = [c for c in all_cols if 'رقم' in c and 'مسلسل' not in c and 'قومي' not in c]

    elif search_type == 'serial':
        # توسيع المرادفات لأعمدة المسلسل لتعمل عبر جميع التبويبات
        serial_candidates = [
            'مسلسل الماكينة', 'مسلسل', 'serial', 'pos serial', 'sn',
            'رقم الماكينة', 'machine serial'
        ]
        target_cols = [
            c for c in all_cols
            if any(tok in str(c).lower() for tok in serial_candidates)
        ]
        # في حال عدم العثور، جرّب أعمدة تحتوي "مسلسل" ككلمة عربية صريحة
        if not target_cols:
            target_cols = [c for c in all_cols if 'مسلسل' in str(c)]
    
    elif search_type == 'machine_code':
         # دعم البحث برقم/كود الماكينة عبر مرادفات متعددة
         mc_tokens = [
             'رقم الماكينة', 'كود الماكينة',
             'machine code', 'machine id', 'pos id', 'terminal id',
             'رقم الجهاز', 'كود الجهاز'
         ]
         mc_tokens = [t.lower() for t in mc_tokens]
         target_cols = [c for c in all_cols if any(tok in str(c).lower() for tok in mc_tokens)]
         # fallback ذكي: أي عمود يحتوي "رقم" مع "ماك" أو يحتوي "code" و"machine"
         if not target_cols:
             target_cols = [c for c in all_cols if ('رقم' in str(c) and 'ماك' in str(c))]
         if not target_cols:
             target_cols = [c for c in all_cols if ('code' in str(c).lower() and 'machine' in str(c).lower())]
        
    elif search_type == 'name':
         # الأولوية للبحث باسم العميل ثم المخبز ثم التاجر
         name_keys_priority = ['اسم العميل', 'اسم المخبز', 'اسم التاجر']
         target_cols = [c for c in name_keys_priority if c in all_cols]
         if not target_cols:
              target_cols = [c for c in all_cols if 'اسم' in c] # Fallback to any 'اسم' column
        
    # إزالة الأعمدة غير الموجودة فعلاً
    target_cols = [col for col in target_cols if col in all_cols]
    # إذا تعذر تحديد أعمدة واضحة للبحث، لا نُوقف العملية
    # بل نستخدم مسار تصفية شامل عبر جميع الأعمدة كحل احتياطي
    use_comprehensive_search = (len(target_cols) == 0)

    # 2. تنفيذ البحث (مسار سريع عبر الفهارس + مسار احتياطي بالتصفية)
    q_norm = _norm_key_text(query).lower()
    hit_indexes: set[int] = set()
    try:
        print(f"[inquiry_debug] start search_type={search_type} q_norm={q_norm} target_cols={target_cols} use_comp={use_comprehensive_search}")
    except Exception:
        pass

    if q_norm:
            if search_type == 'code':
                # مطابقة دقيقة أولاً
                exact_hits = indexes.get('code', {}).get(q_norm, [])
                if exact_hits:
                    hit_indexes.update(exact_hits)
                    try:
                        print(f"[inquiry_debug] path=index_exact_code q={q_norm} hits={len(exact_hits)}")
                    except Exception:
                        pass
            elif search_type == 'serial':
                # تفضيل المطابقة التامة؛ استخدام البادئات فقط إذا لم توجد مطابقة تامة
                exact_hits = indexes.get('serial', {}).get(q_norm, [])
                if exact_hits:
                    hit_indexes.update(exact_hits)
                    try:
                        print(f"[inquiry_debug] path=index_exact_serial q={q_norm} hits={len(exact_hits)}")
                    except Exception:
                        pass
                else:
                    if len(q_norm) >= 5:
                        pref5 = indexes.get('serial_prefix5', {}).get(q_norm[:5], [])
                        hit_indexes.update(pref5)
                        try:
                            print(f"[inquiry_debug] path=index_prefix_serial_5 q={q_norm[:5]} hits={len(pref5)}")
                        except Exception:
                            pass
                    if len(q_norm) >= 3:
                        pref3 = indexes.get('serial_prefix3', {}).get(q_norm[:3], [])
                        hit_indexes.update(pref3)
                        try:
                            print(f"[inquiry_debug] path=index_prefix_serial_3 q={q_norm[:3]} hits={len(pref3)}")
                        except Exception:
                            pass
            elif search_type == 'machine_code':
                # تفضيل المطابقة التامة؛ استخدام البادئات فقط إذا غابت المطابقة التامة
                exact_hits = indexes.get('machine_code', {}).get(q_norm, [])
                if exact_hits:
                    hit_indexes.update(exact_hits)
                    try:
                        print(f"[inquiry_debug] path=index_exact_machine_code q={q_norm} hits={len(exact_hits)}")
                    except Exception:
                        pass
                else:
                    if len(q_norm) >= 5:
                        pref5 = indexes.get('machine_code_prefix5', {}).get(q_norm[:5], [])
                        hit_indexes.update(pref5)
                        try:
                            print(f"[inquiry_debug] path=index_prefix_machine_code_5 q={q_norm[:5]} hits={len(pref5)}")
                        except Exception:
                            pass
                    if len(q_norm) >= 3:
                        pref3 = indexes.get('machine_code_prefix3', {}).get(q_norm[:3], [])
                        hit_indexes.update(pref3)
                        try:
                            print(f"[inquiry_debug] path=index_prefix_machine_code_3 q={q_norm[:3]} hits={len(pref3)}")
                        except Exception:
                            pass
            elif search_type == 'name':
                # الاسم الكامل أولاً
                name_hits = indexes.get('name', {}).get(q_norm, [])
                hit_indexes.update(name_hits)
                try:
                    print(f"[inquiry_debug] path=index_exact_name q={q_norm} hits={len(name_hits)}")
                except Exception:
                    pass
                # كلمات الاسم (مطابقة كاملة)
                for tok in [t for t in q_norm.split(' ') if t]:
                    tok_hits = indexes.get('name_token', {}).get(tok, [])
                    hit_indexes.update(tok_hits)
                    # بادئة للكلمات
                    if len(tok) >= 5:
                        pref5 = indexes.get('name_token_prefix5', {}).get(tok[:5], [])
                        hit_indexes.update(pref5)
                    if len(tok) >= 3:
                        pref3 = indexes.get('name_token_prefix3', {}).get(tok[:3], [])
                        hit_indexes.update(pref3)
    
    
    filtered_df = None
    # عند البحث بنوع 'code' نستخدم مطابقة contains عبر الأعمدة المحددة
    if search_type == 'code' and q_norm:
        try:
            # استخدم دالة التصفية الموحدة للبحث الجزئي داخل الأعمدة المستهدفة
            filtered_df = _filter_dataframe(mapped_df, query, search_cols=(target_cols or []))
            try:
                print(f"[inquiry_debug] path=code_contains_target_cols q={query} cols={target_cols} count={len(filtered_df) if filtered_df is not None else 0}")
            except Exception:
                pass
        except Exception:
            filtered_df = mapped_df.iloc[0:0]
    if (filtered_df is None) and hit_indexes:
        try:
            filtered_df = mapped_df.loc[sorted(hit_indexes)]
        except Exception:
            # في حال فشل التقطيع بالمؤشرات، نستخدم الإسقاط عبر mask
            mask = mapped_df.index.isin(list(hit_indexes))
            filtered_df = mapped_df[mask]
        
    elif filtered_df is None:
        # تصفية ذكية عبر الأعمدة المستهدفة؛ وإذا لم تُحدد أعمدة، استخدم البحث الشامل عبر كل الأعمدة
        filtered_df = _filter_dataframe(mapped_df, query, search_cols=(None if use_comprehensive_search else target_cols))
        try:
            path = 'comprehensive_contains' if use_comprehensive_search else 'target_cols_contains'
            print(f"[inquiry_debug] path={path} q={query} cols={None if use_comprehensive_search else target_cols} count={len(filtered_df)}")
        except Exception:
            pass

    filtered_df = filtered_df.copy()

    if filtered_df.empty:
        try:
            print(f"[inquiry_debug] empty_result q={query} search_type={search_type}")
        except Exception:
            pass
        return {'success': False, 'message': f'لم يتم العثور على نتائج للبحث عن "{query}".', 'items': []}

    
    filtered_df = _drop_empty_columns(filtered_df) 

    # 4. التجميع الأصلي للبيانات
    grouped_nested_results = _group_search_results(filtered_df, category)
    
    total_found_entities = len(grouped_nested_results)
    
    if not grouped_nested_results:
        try:
            print(f"[inquiry_debug] grouped_empty q={query} filtered_count={len(filtered_df)}")
        except Exception:
            pass
        return {'success': False, 'message': 'تم العثور على سجلات، لكن لم يتم تجميعها في كيان صالح (برجاء التحقق من أعمدة رقم العميل/اسم العميل).', 'items': []}

    # 5. إعادة هيكلة النتيجة للواجهة (نستخدم النتيجة الأولى فقط)
    first_group = grouped_nested_results[0]
    common_data = first_group['common_data']
    
    # تحديد مفاتيح الكيان
    group_keys = first_group.get('group_keys', ['رقم العميل', 'اسم العميل'])
    customer_code = common_data.get(group_keys[0], '-')
    customer_name = common_data.get(group_keys[1], '-')

    # 💡 6. استخراج الحقول الديناميكية (من ملف all data.xlsx)
    # أسماء الأعمدة المتوقعة من ملف all data.xlsx
    DYNAMIC_FIELDS_COLS = {
        'ماكينة فرع': 'لدية ماكينة فرع',
        'ماكينه فرع': 'لدية ماكينة فرع',
        'إجمالي الحوالات': 'إجمالي الحوالات',
        'الحواله': 'إجمالي الحوالات',
        'اخر وضع للماكينة': 'اخر وضع للماكينة',
        'الحاله': 'اخر وضع للماكينة',
        'اخر تاريخ سفر للماكينة': 'اخر تاريخ سفر للماكينة',
        'تاريخ السفر للصيانه': 'اخر تاريخ سفر للماكينة',
        'تاريخ السفر للصيانة': 'اخر تاريخ سفر للماكينة',
    }
    
    dynamic_fields = {}
    
    # استخراج الحقول الديناميكية
    for col_in_data, key_for_display in DYNAMIC_FIELDS_COLS.items():
        dynamic_fields[key_for_display] = common_data.get(col_in_data, '-')
        # إزالة العمود من البيانات المشتركة بعد استخلاصه
        if col_in_data in common_data:
             del common_data[col_in_data] 

    # تحميل الماكينات الأساسية من خدمات التجار (trader_primary) ومطابقة الكود/الاسم
    def _load_primary_df() -> pd.DataFrame:
        try:
            row = _load_state("trader_primary")
            if not row or not row.data_json:
                return pd.DataFrame()
            dfp = _json_to_df(row.data_json)
            map_row = _load_state("trader_primary:__mapping__")
            mapping = json.loads(map_row.mapping_json) if (map_row and map_row.mapping_json) else {}
            dfp = _apply_mapping(dfp, mapping)
            return _drop_empty_columns(dfp)
        except Exception:
            return pd.DataFrame()

    def _detect_primary_key_cols(cols: list[str]) -> tuple[str|None, str|None]:
        """تحديد أعمدة الكود والاسم بذكاء من أسماء الأعمدة الفعلية.
        يعتمد على مجموعة واسعة من المرادفات العربية والإنجليزية ويستخدم مطابقة تحتوي على كلمات رئيسية.
        """
        if not cols:
            return None, None

        all_cols_norm = [str(c).strip() for c in cols]

        # مفردات أساسية
        code_tokens = ['رقم', 'الكود', 'id', 'code']
        name_tokens = ['اسم', 'name']
        entity_tokens = ['عميل', 'تاجر', 'مخبز', 'فرع', 'customer', 'trader', 'bakery', 'branch']

        # مرشّحات مرجّحة صريحة كاملة أولاً
        explicit_code_candidates = [
            'رقم العميل','Customer Code','Customer_ID','Customer ID','رقم التاجر','trader_id','رقم المخبز','Bakery ID'
        ]
        explicit_name_candidates = [
            'اسم العميل','اسم المخبز','اسم التاجر','Customer Name','Trader Name','trader_name','Bakery Name'
        ]

        def pick_explicit(cands):
            for c in cands:
                if c in all_cols_norm:
                    return c
            return None

        code_col = pick_explicit(explicit_code_candidates)
        name_col = pick_explicit(explicit_name_candidates)

        # إذا لم يُكتشف صراحةً، استخدم مطابقة تعتمد على احتواء الكلمات الرئيسية
        def contains_any(haystack: str, needles: list[str]) -> bool:
            h = haystack.lower()
            return any(n in h for n in needles)

        if code_col is None:
            # ابحث عن عمود يحتوي كلمات (رقم|الكود|id|code) مع كيان (عميل|مخبز|تاجر|فرع)
            ranked = []
            for c in all_cols_norm:
                score = 0
                if contains_any(c, code_tokens):
                    score += 2
                if contains_any(c, entity_tokens):
                    score += 1
                if score > 0:
                    ranked.append((score, c))
            if ranked:
                ranked.sort(key=lambda x: (-x[0], len(x[1])))
                code_col = ranked[0][1]

        if name_col is None:
            ranked = []
            for c in all_cols_norm:
                score = 0
                if contains_any(c, name_tokens):
                    score += 2
                if contains_any(c, entity_tokens):
                    score += 1
                if score > 0:
                    ranked.append((score, c))
            if ranked:
                ranked.sort(key=lambda x: (-x[0], len(x[1])))
                name_col = ranked[0][1]

        # Fallback أخير: أي عمود يحتوي "رقم" للكود و"اسم" للاسم
        if code_col is None:
            for c in all_cols_norm:
                cl = c.lower()
                if ('رقم' in cl) or ('الكود' in cl) or ('id' in cl) or ('code' in cl):
                    code_col = c
                    break
        if name_col is None:
            for c in all_cols_norm:
                cl = c.lower()
                if ('اسم' in cl) or ('name' in cl):
                    name_col = c
                    break

        return code_col, name_col

    PRIMARY_FIELD_CANDS = {
        'لدية ماكينة فرع': [
            'لدية ماكينة فرع', 'لديه ماكينة فرع', 'ماكينة فرع', 'ماكينه فرع', 'Has Branch Machine', 'Branch Machine', 'يوجد ماكينة فرع'
        ],
        'إجمالي الحوالات': [
            'إجمالي الحوالات', 'اجمالي الحوالات', 'الحواله', 'الحوالات', 'Total Transfers', 'Total Transfer', 'Transfers', 'Transfer'
        ],
        'اخر وضع للماكينة': [
            'اخر وضع للماكينة', 'آخر وضع للماكينة', 'اخر وضع', 'آخر وضع', 'الحاله', 'حالة الماكينة', 'Last Status', 'Status'
        ],
        'اخر تاريخ سفر للماكينة': [
            'اخر تاريخ سفر للماكينة', 'آخر تاريخ سفر للماكينة', 'اخر تاريخ سفر', 'آخر تاريخ سفر', 'تاريخ السفر للصيانه', 'تاريخ السفر للصيانة', 'Last Travel Date', 'Travel Date for Maintenance', 'Travel Date'
        ],
        'القائم بتسليم ماكينة الفرع': [
            'القائم بتسليم ماكينة الفرع', 'القائم بالتسليم', 'مسؤول التسليم', 'مسئول التسليم', 'Responsible for Branch Machine Delivery'
        ],
    }

    primary_debug = {}
    primary_record = {}
    primary_match_mode = 'none'
    _primary_df = _load_primary_df()
    if not _primary_df.empty:
        code_col, name_col = _detect_primary_key_cols(list(_primary_df.columns))
        code_norm = _norm_key_text(str(customer_code)) if _textify(customer_code) != '' else None
        name_norm = _norm_key_text(str(customer_name)) if _textify(customer_name) != '' else None
        mask = None
        if code_col and code_norm is not None:
            mask = (_primary_df[code_col].astype(str).apply(_norm_key_text) == code_norm)
        if name_col and name_norm is not None:
            nm = (_primary_df[name_col].astype(str).apply(_norm_key_text) == name_norm)
            mask = (mask & nm) if mask is not None else nm
        filtered_primary = _primary_df[mask] if mask is not None else pd.DataFrame()
        primary_debug = {
            'source_cols': list(_primary_df.columns),
            'match_cols': {'code': code_col, 'name': name_col},
            'match_rows': int(filtered_primary.shape[0]) if not filtered_primary.empty else 0,
        }
        if not filtered_primary.empty:
            primary_match_mode = 'intersection'
        # إن لم نجد صفًا باستخدام التقاطع، جرّب الاتحاد (مطابقة بالرقم أو الاسم)
        if (filtered_primary.empty) and (code_col or name_col):
            union_mask = None
            if code_col and code_norm is not None:
                union_mask = (_primary_df[code_col].astype(str).apply(_norm_key_text) == code_norm)
            if name_col and name_norm is not None:
                nm2 = (_primary_df[name_col].astype(str).apply(_norm_key_text) == name_norm)
                union_mask = (union_mask | nm2) if union_mask is not None else nm2
            if union_mask is not None:
                filtered_primary = _primary_df[union_mask]
                primary_debug['match_rows_union'] = int(filtered_primary.shape[0])
                primary_debug['used_union'] = True
                if not filtered_primary.empty:
                    primary_match_mode = 'union'
        if not filtered_primary.empty:
            # اختيار الصف الأفضل إذا تعددت الصفوف في الاتحاد
            try:
                best_idx = 0
                if filtered_primary.shape[0] > 1:
                    # حاول مطابقة الاسم بدقة إذا توفر
                    if name_col and name_norm is not None and name_col in filtered_primary.columns:
                        exact_name = filtered_primary[name_col].astype(str).apply(_norm_key_text) == name_norm
                        match_idxs = list(filtered_primary[exact_name].index)
                        if match_idxs:
                            best_idx = match_idxs[0]
                    # أو مطابقة الكود بدقة إذا توفر
                    elif code_col and code_norm is not None and code_col in filtered_primary.columns:
                        exact_code = filtered_primary[code_col].astype(str).apply(_norm_key_text) == code_norm
                        match_idxs = list(filtered_primary[exact_code].index)
                        if match_idxs:
                            best_idx = match_idxs[0]
                first_row = filtered_primary.loc[best_idx].to_dict()
            except Exception:
                first_row = filtered_primary.iloc[0].to_dict()
            # حفظ الصف الكامل لاستخدامه في الواجهة
            try:
                primary_record = {k: ('' if _textify(v) == '' else v) for k, v in first_row.items()}
            except Exception:
                primary_record = first_row
            # Override قيم الحقول الديناميكية من مصدر الماكينات الأساسية إذا توفرت
            for display_key, cand_list in PRIMARY_FIELD_CANDS.items():
                val = None
                for cand in cand_list:
                    if cand in first_row and _textify(first_row.get(cand)) != '':
                        val = first_row.get(cand)
                        break
                if val is not None:
                    dynamic_fields[display_key] = val

    # بناء جزء ماكينات الفرع والحوالات من صف All Data المطابق تقاطعاً فقط
    branch_section = {
        'ماكينه فرع': '-',
        'الحواله': '-',
        'الحاله': '-',
        'تاريخ السفر للصيانه': '-',
        'القائم بتسليم ماكينة الفرع': '-',
    }
    try:
        if primary_record:
            # استخدم المرادفات للعثور على القيم المطلوبة من الصف (سواء تقاطع أو اتحاد)
            field_synonyms = {
                'ماكينه فرع': ['ماكينه فرع', 'ماكينة فرع', 'لدية ماكينة فرع', 'لديه ماكينة فرع', 'Has Branch Machine', 'Branch Machine', 'يوجد ماكينة فرع'],
                'الحواله': ['الحواله', 'الحوالات', 'إجمالي الحوالات', 'Total Transfers', 'Transfers'],
                'الحاله': ['الحاله', 'حالة الماكينة', 'اخر وضع للماكينة', 'آخر وضع للماكينة', 'Last Status', 'Status'],
                'تاريخ السفر للصيانه': ['تاريخ السفر للصيانه', 'تاريخ السفر للصيانة', 'اخر تاريخ سفر للماكينة', 'آخر تاريخ سفر للماكينة', 'آخر تاريخ سفر', 'Last Travel Date'],
                'القائم بتسليم ماكينة الفرع': ['القائم بتسليم ماكينة الفرع', 'القائم بالتسليم', 'مسؤول التسليم', 'مسئول التسليم']
            }
            for display_key, cands in field_synonyms.items():
                for c in cands:
                    if c in primary_record and _textify(primary_record.get(c)) != '':
                        branch_section[display_key] = primary_record.get(c)
                        break
    except Exception:
        pass
    
    # 7. تجهيز بيانات المسلسلات مبكرًا لاستخدامها في تصفية سجل الزيارات
    
    # 💡 تم إزالة دالة convert_primary_secondary ومنطق تطبيقها، حيث أن التحويل أصبح يتم
    # مباشرة بعد الاستيراد في دالة _merge_all.
    
    serial_list = []
    # 💡 خريطة لأسماء أعمدة المسلسلات
    SERIAL_MAP = {
        'مسلسل الماكينة': 'مسلسل',
        'رقم الماكينة': 'رقم الماكينة',
        'ماكينة رئيسية/فرعية': 'رئيسية/فرعية', 
        'شريحة 1': 'شريحة1',
        'شريحة 2': 'شريحة2',
        'حالة الماكينة': 'حالة الماكينة',
        'حالة نظام المطحن': 'حالة نظام المطحن', 
        'SW_AC_SUP': 'SW_AC_SUP', 
        'SW_IC_SUP': 'SW_IC_SUP',
        'SW_OD_SUP': 'SW_OD_SUP',
        'POS_VERSION': 'POS_VERSION', 
        'اسم الخبز': 'اسم الخبز',
        'LOAF_BALANCE1': 'LOAF_BALANCE1',
        'ساعة بدء البيع': 'ساعة بدء البيع',
        'ساعة نهاية البيع': 'ساعة نهاية البيع',
    }
    
    for machine in first_group['machine_details']:
        serial_item = {}
        for original_col, new_key in SERIAL_MAP.items():
            value = machine.get(original_col, '-')
            
            # 💡 بما أن التحويل تم عند الاستيراد، نكتفي بأخذ القيمة مباشرة
            serial_item[new_key] = value
            
        serial_list.append(serial_item)
        
    # 9. تجهيز بيانات العميل الأساسية بالترتيب المطلوب
    
    # 💡 التعديل: خريطة الأعمدة للبحث عن بدائل في البيانات المستوردة (لإصلاح رقم المحمول واسم المسؤل)
    CUSTOMER_FETCH_MAP = {
        'رقم المحمول': ['رقم المحمول', 'المحمول', 'موبايل', 'رقم الهاتف المحمول'],
        'اسم المسؤل': ['اسم المسؤل', 'اسم المسئول', 'المسئول', 'المسؤول', 'مسئول'], 
        
        'رقم هاتف': ['رقم هاتف', 'تليفون', 'هاتف', 'ثابت'],
        'الرقم القومي': ['الرقم القومي', 'بطاقة', 'رقم بطاقة'],
        'الادارة': ['الادارة', 'المديرية'],
        'المكتب': ['المكتب', 'الشعبة'],
        'رقم العميل': ['رقم العميل', 'كود العميل'],
        'اسم العميل': ['اسم العميل', 'اسم المخبز', 'اسم التاجر'],
        'حالة العميل': ['حالة العميل', 'حالة'],
        'ملاحظات': ['ملاحظات', 'Note'],
    }
    
    customer_data = {}
    for display_key in CUSTOMER_DETAIL_COLS:
        value = '-'
        # يتم استخدام مفتاح العرض أولاً، وإذا لم يكن موجودًا، يتم البحث في البدائل
        potential_keys = CUSTOMER_FETCH_MAP.get(display_key, [display_key])
        
        for p_key in potential_keys:
            # التحقق من وجود المفتاح في البيانات المشتركة وأن قيمته ليست فارغة
            if p_key in common_data and _textify(common_data.get(p_key)) != '':
                value = common_data.get(p_key)
                break
                
        customer_data[display_key] = value

    # 10. الرسالة النهائية
    message = f'تم العثور على {total_found_entities} سجل(ات) كيان مطابق. (العميل: {customer_name})'
    if total_found_entities > 1:
        message += ' يمكنك التنقل بين الكيانات باستخدام أزرار التنقل أو مفتاحي السهمين (↑ و ↓).'
    
    # 9. تحميل سجل الزيارات الفعلي للسنة/الشهر الحاليين من خدمات التجار (المترددين)
    def _load_visit_history_df() -> tuple[pd.DataFrame, dict]:
        global _VISIT_CACHE
        # ملاحظة: بيانات خدمات التجار تُخزن عالمياً بدون user_id
        # سنجمّع بيانات السنة الحالية من جميع الفترات المتاحة (label="YYYY" أو "YYYY-MM")
        try:
            now = datetime.now()
            year = now.year
            month = now.month
            month_label = f"{year}-{month:02d}"

            # استخدام كاش خفيف لتقليل القراءة المتكررة لنفس الفترة
            try:
                if (_VISIT_CACHE.get('df') is not None) and (time() - (_VISIT_CACHE.get('ts') or 0) < _VISIT_CACHE_TTL_SEC):
                    cached_df = _VISIT_CACHE['df']
                    cached_meta = _VISIT_CACHE['meta']
                    return cached_df.copy(), (dict(cached_meta) if isinstance(cached_meta, dict) else cached_meta)
            except Exception:
                pass

            # فرض استخدام بيانات البرنامج الحديثة أولاً دائمًا
            try:
                rp_rows = (ReportState.query
                           .filter(ReportState.category == 'trader_frequent:recent_program')
                           .order_by(ReportState.created_at.desc())
                           .all())
                map_row = ReportState.query.filter_by(category="trader_frequent:__mapping__").first()
                mapping = json.loads(map_row.mapping_json) if (map_row and map_row.mapping_json) else {}
                df = pd.DataFrame()
                for r in rp_rows:
                    if not r.data_json:
                        continue
                    d = _json_to_df(r.data_json)
                    if not d.empty:
                        d = _apply_mapping(d, mapping)
                        try:
                            d['_الفترة'] = 'البيانات الحديثة (البرنامج)'
                        except Exception:
                            pass
                        df = pd.concat([df, d], ignore_index=True) if not df.empty else d
                result_df = _drop_empty_columns(df)
                meta = {'month_label': None, 'year_label': None, 'recent_program': True, 'source': 'recent_program'}
                try:
                    _VISIT_CACHE = {'df': result_df.copy(), 'meta': meta.copy(), 'ts': time()}
                except Exception:
                    pass
                return result_df, meta
            except Exception:
                # إذا فشلت القراءة الحديثة، نكمل بمنطق الشهر/السنة
                pass

            # استعلامات مستهدفة بدل مسح جميع الفترات
            exact_month_row = (ReportState.query
                               .filter(ReportState.category == f"trader_frequent:{month_label}")
                               .order_by(ReportState.created_at.desc())
                               .first())
            year_rows = (ReportState.query
                         .filter(ReportState.category == f"trader_frequent:{year}")
                         .order_by(ReportState.created_at.desc())
                         .all())
            months_dash = (ReportState.query
                           .filter(ReportState.category.like(f"trader_frequent:{year}-%"))
                           .order_by(ReportState.created_at.desc())
                           .all())
            months_slash = (ReportState.query
                            .filter(ReportState.category.like(f"trader_frequent:{year}/%"))
                            .order_by(ReportState.created_at.desc())
                            .all())
            months_of_year_rows = months_dash + months_slash
            fallback_month_row = (ReportState.query
                                  .filter(ReportState.category.like("trader_frequent:%-%"))
                                  .order_by(ReportState.created_at.desc())
                                  .first())
            # تعيين مخطط إعادة التسمية لترتيب/إعادة تسمية الأعمدة (إن وُجد)
            map_row = ReportState.query.filter_by(category="trader_frequent:__mapping__").first()
            mapping = json.loads(map_row.mapping_json) if (map_row and map_row.mapping_json) else {}

            # انتقاء فترات السنة الحالية
            def _label_of(cat: str) -> str:
                try:
                    return cat.split(":", 1)[1]
                except Exception:
                    return ""

            # تم توليد year_rows و months_of_year_rows و exact/fallback عبر استعلامات مستهدفة أعلاه

            # 1) تجميع السنة: إما من ملف سنة واحدة أو من شهور السنة
            df_year = pd.DataFrame()
            if year_rows:
                for r in year_rows:
                    d = _json_to_df(r.data_json)
                    if not d.empty:
                        d = _apply_mapping(d, mapping)
                        try:
                            d['_الفترة'] = str(year)
                        except Exception:
                            pass
                        df_year = pd.concat([df_year, d], ignore_index=True) if not df_year.empty else d
            elif months_of_year_rows:
                for r in months_of_year_rows:
                    d = _json_to_df(r.data_json)
                    if not d.empty:
                        d = _apply_mapping(d, mapping)
                        try:
                            lab = _label_of(r.category)
                            d['_الفترة'] = lab.replace('/', '-')
                        except Exception:
                            pass
                        df_year = pd.concat([df_year, d], ignore_index=True) if not df_year.empty else d

            # 2) تحديد ملف الشهر: مطابق للشهر الحالي أو أحدث ملف شهري متاح
            selected_month_row = exact_month_row or fallback_month_row
            selected_month_label = None
            df_month = pd.DataFrame()
            if selected_month_row is not None:
                selected_month_label = _label_of(selected_month_row.category).replace('/', '-')
                d = _json_to_df(selected_month_row.data_json)
                if not d.empty:
                    d = _apply_mapping(d, mapping)
                    try:
                        d['_الفترة'] = selected_month_label
                    except Exception:
                        pass
                    df_month = d
            # حساب مجموعة العضوية من بيانات البرنامج الحديثة (recent_program)
            code_set = set()
            serial_set = set()
            try:
                rp_rows_for_sets = (ReportState.query
                                    .filter(ReportState.category == 'trader_frequent:recent_program')
                                    .order_by(ReportState.created_at.desc())
                                    .all())
                rp_df_sets = pd.DataFrame()
                for rr in rp_rows_for_sets:
                    if not rr.data_json:
                        continue
                    dd = _json_to_df(rr.data_json)
                    if not dd.empty:
                        dd = _apply_mapping(dd, mapping)
                        rp_df_sets = pd.concat([rp_df_sets, dd], ignore_index=True) if not rp_df_sets.empty else dd
                if not rp_df_sets.empty:
                    std_rp = _standardize_visit_df(rp_df_sets)
                    if 'رقم العميل' in std_rp.columns:
                        code_set = set(std_rp['رقم العميل'].astype(str).apply(_norm_key_text))
                    if 'مسلسل' in std_rp.columns:
                        serial_set = set(std_rp['مسلسل'].astype(str).apply(_norm_key_text))
            except Exception:
                pass

            # اختيار المصدر تلقائياً:
            # - إذا توفر ملف السنة الحالية: نستخدمه فقط ونصنف تحت "السنة الحالية"
            # - وإلا إذا توفر ملف الشهر الحالي/أحدث شهر: نستخدمه فقط ونصنف تحت "الشهر الحالي"
            # - وإلا نستخدم بيانات البرنامج الحديثة ونصنفها تحت "الشهر الحالي"

            if not df_year.empty:
                # فلترة السنة وفق عضوية recent_program إن توفرت مجموعات عضوية
                if code_set or serial_set:
                    std_y = _standardize_visit_df(df_year)
                    mask_code = std_y['رقم العميل'].astype(str).apply(_norm_key_text).isin(code_set) if ('رقم العميل' in std_y.columns and code_set) else None
                    mask_serial = std_y['مسلسل'].astype(str).apply(_norm_key_text).isin(serial_set) if ('مسلسل' in std_y.columns and serial_set) else None
                    if mask_code is not None and mask_serial is not None:
                        year_mask = (mask_code | mask_serial)
                    elif mask_code is not None:
                        year_mask = mask_code
                    elif mask_serial is not None:
                        year_mask = mask_serial
                    else:
                        year_mask = pd.Series([True] * std_y.shape[0])
                    df_year = df_year[year_mask.values]
                result_df = _drop_empty_columns(df_year)
                meta = {'month_label': None, 'year_label': str(year), 'source': 'year'}
                try:
                    _VISIT_CACHE = {'df': result_df.copy(), 'meta': meta.copy(), 'ts': time()}
                except Exception:
                    pass
                return result_df, meta
            if not df_month.empty:
                # فلترة الشهر وفق عضوية recent_program إن توفرت مجموعات عضوية
                if code_set or serial_set:
                    std_m = _standardize_visit_df(df_month)
                    mask_code = std_m['رقم العميل'].astype(str).apply(_norm_key_text).isin(code_set) if ('رقم العميل' in std_m.columns and code_set) else None
                    mask_serial = std_m['مسلسل'].astype(str).apply(_norm_key_text).isin(serial_set) if ('مسلسل' in std_m.columns and serial_set) else None
                    if mask_code is not None and mask_serial is not None:
                        month_mask = (mask_code | mask_serial)
                    elif mask_code is not None:
                        month_mask = mask_code
                    elif mask_serial is not None:
                        month_mask = mask_serial
                    else:
                        month_mask = pd.Series([True] * std_m.shape[0])
                    df_month = df_month[month_mask.values]
                result_df = _drop_empty_columns(df_month)
                meta = {'month_label': selected_month_label, 'year_label': None, 'source': 'month'}
                try:
                    _VISIT_CACHE = {'df': result_df.copy(), 'meta': meta.copy(), 'ts': time()}
                except Exception:
                    pass
                return result_df, meta

            # بيانات البرنامج الحديثة كحل احتياطي
            try:
                rp_rows = (ReportState.query
                           .filter(ReportState.category == 'trader_frequent:recent_program')
                           .order_by(ReportState.created_at.desc())
                           .all())
                map_row = ReportState.query.filter_by(category="trader_frequent:__mapping__").first()
                mapping = json.loads(map_row.mapping_json) if (map_row and map_row.mapping_json) else {}
                df = pd.DataFrame()
                for r in rp_rows:
                    if not r.data_json:
                        continue
                    d = _json_to_df(r.data_json)
                    if not d.empty:
                        d = _apply_mapping(d, mapping)
                        try:
                            d['_الفترة'] = 'البيانات الحديثة (البرنامج)'
                        except Exception:
                            pass
                        df = pd.concat([df, d], ignore_index=True) if not df.empty else d
                result_df = _drop_empty_columns(df)
                meta = {'month_label': None, 'year_label': None, 'recent_program': True, 'source': 'recent_program'}
                try:
                    _VISIT_CACHE = {'df': result_df.copy(), 'meta': meta.copy(), 'ts': time()}
                except Exception:
                    pass
                return result_df, meta
            except Exception:
                return pd.DataFrame(), {'month_label': None, 'year_label': None, 'source': 'month'}
        except Exception:
            # في أي خطأ، نعود لإستراتيجية قديمة (إن وُجدت) أو نُرجع فارغ
            try:
                row_vis = _load_state("visit_history")
                if not row_vis or not row_vis.data_json:
                    return pd.DataFrame(), {'month_label': None, 'year_label': None, 'source': 'month'}
                df_vis = _json_to_df(row_vis.data_json)
                map_row = _load_state("visit_history:__mapping__")
                mapping = json.loads(map_row.mapping_json) if (map_row and map_row.mapping_json) else {}
                df_vis = _apply_mapping(df_vis, mapping)
                return _drop_empty_columns(df_vis), {'month_label': None, 'year_label': None, 'source': 'month'}
            except Exception:
                return pd.DataFrame(), {'month_label': None, 'year_label': None, 'source': 'month'}

    def _detect_visit_columns(cols: list[str]) -> tuple[str|None, str|None, str|None, str|None]:
        # اكتشاف أعمدة: مسلسل، التاريخ، رقم العميل، اسم العميل (اختيارية)
        serial_candidates = ['مسلسل', 'مسلسل الماكينة', 'Serial', 'POS Serial', 'POS', 'رقم الماكينة']
        date_candidates = ['التاريخ', 'تاريخ الزيارة', 'Date', 'Visit Date', 'تاريخ']
        customer_code_candidates = ['رقم العميل', 'Customer Code', 'Customer_ID', 'client_code', 'trader_id', 'code', 'ID', 'id', 'Bakery ID', 'رقم المخبز', 'رقم التاجر']
        customer_name_candidates = ['اسم العميل', 'اسم المخبز', 'اسم التاجر', 'Customer Name', 'Trader Name', 'Bakery Name', 'trader_name', 'name', 'customer_name', 'trader_name_ar']
        def pick(cands):
            for c in cands:
                if c in cols:
                    return c
            return None
        return (
            pick(serial_candidates),
            pick(date_candidates),
            pick(customer_code_candidates),
            pick(customer_name_candidates)
        )

    visit_history_df, _period_meta = _load_visit_history_df()

    # جمع مسلسلات المجموعة الحالية لاستخدامها في التصفية إذا لم يتوفر رقم العميل في سجل الزيارات
    current_serials = set()
    try:
        for md in first_group.get('machine_details', []):
            if 'مسلسل الماكينة' in md and _textify(md.get('مسلسل الماكينة')) != '':
                current_serials.add(_textify(md.get('مسلسل الماكينة')))
    except Exception:
        current_serials = set()

    visit_debug = {}
    # إضافة معلومات مطابقة الماكينات الأساسية للوحة التشخيص
    try:
        visit_debug['primary'] = primary_debug
    except Exception:
        pass
    if not visit_history_df.empty:
        # إعادة بناء المنطق: توحيد الأعمدة ثم مطابقة صارمة ومرنة
        dfv = _standardize_visit_df(visit_history_df)
        visit_debug['source_cols'] = list(visit_history_df.columns)
        visit_debug['std_cols'] = list(dfv.columns)
        visit_debug['source_rows'] = int(visit_history_df.shape[0])
        code_norm = _norm_key_text(str(customer_code)) if _textify(customer_code) != '' else None
        name_norm = _norm_key_text(str(customer_name)) if _textify(customer_name) != '' else None
        serials_norm = set(_norm_key_text(s) for s in current_serials) if current_serials else set()
        # فلترة حسب النوع المطلوب (مخابز/تموين/استبدال) قبل بناء الأقنعة
        pre_dfv = dfv
        target_type = CATEGORIES.get(category, category)
        type_norm = _norm_key_text(str(target_type)) if _textify(target_type) != '' else None
        type_mask = None
        if ('النوع' in pre_dfv.columns) and type_norm is not None:
            try:
                type_mask = (pre_dfv['النوع'].apply(_textify).apply(_norm_key_text) == type_norm)
            except Exception:
                type_mask = None
        if type_mask is not None:
            pre_dfv = pre_dfv[type_mask]
        visit_debug['type_norm'] = type_norm
        visit_debug['type_rows'] = int(pre_dfv.shape[0])

        mask = None
        # استخدام _textify قبل _norm_key_text لضمان إزالة الأجزاء مثل ".0" وتحويل الأرقام إلى نص موحد
        if 'رقم العميل' in dfv.columns and code_norm is not None:
            mask = (dfv['رقم العميل'].apply(_textify).apply(_norm_key_text) == code_norm)
        if 'اسم العميل' in dfv.columns and name_norm is not None:
            nm = (dfv['اسم العميل'].apply(_textify).apply(_norm_key_text) == name_norm)
            mask = (mask & nm) if mask is not None else nm
        # طبّق التقاطع إن توفر كلا العمودين
        if mask is not None:
            dfv = dfv[mask]
            # بعد تطبيق التقاطع، طبّق فلترة النوع إن وُجدت
            if type_mask is not None:
                dfv = dfv[type_mask]
        # إن لم تُنتج نتيجة، جرّب اتحاد الشرطين إن كانا متوفرين
        if dfv.empty and ('رقم العميل' in pre_dfv.columns or 'اسم العميل' in pre_dfv.columns):
            union_mask = None
            if 'رقم العميل' in pre_dfv.columns and code_norm is not None:
                union_mask = (pre_dfv['رقم العميل'].apply(_textify).apply(_norm_key_text) == code_norm)
            if 'اسم العميل' in pre_dfv.columns and name_norm is not None:
                nm = (pre_dfv['اسم العميل'].apply(_textify).apply(_norm_key_text) == name_norm)
                union_mask = (union_mask | nm) if union_mask is not None else nm
            if union_mask is not None:
                dfv = pre_dfv[union_mask]
        # وإن بقيت فارغة، اعتمد المسلسلات كحل أخير
        if dfv.empty and ('مسلسل' in pre_dfv.columns) and serials_norm:
            dfv = pre_dfv[pre_dfv['مسلسل'].apply(_textify).apply(_norm_key_text).isin(serials_norm)]
        # سجّل عدّ مفصول للرقم وللمسلسل مع فلترة النوع
        try:
            code_count = 0
            if ('رقم العميل' in pre_dfv.columns) and (code_norm is not None):
                code_count = int((pre_dfv['رقم العميل'].apply(_textify).apply(_norm_key_text) == code_norm).sum())
            serial_count = 0
            if ('مسلسل' in pre_dfv.columns) and serials_norm:
                serial_count = int(pre_dfv['مسلسل'].apply(_textify).apply(_norm_key_text).isin(serials_norm).sum())
            visit_debug['code_count'] = code_count
            visit_debug['serial_count'] = serial_count
        except Exception:
            pass
        visit_debug['code_norm'] = code_norm
        visit_debug['name_norm'] = name_norm
        visit_debug['serials_norm_count'] = len(serials_norm)
        visit_debug['matched_rows'] = int(dfv.shape[0])
        visit_debug['count_mode'] = 'textify_norm_matching'
        # ملاحظة: سنمرر إطار البيانات الموحّد بالكامل إلى دالة العد التي تتعامل مع وجود/غياب التاريخ والمسلسل.
        # تحديد الفترة تلقائياً من الميتاداتا: السنة ← year، الحديثة ← recent_program، وإلا شهر
        if isinstance(_period_meta, dict) and _period_meta.get('source') in ['year','month','recent_program']:
            if _period_meta.get('source') == 'year':
                visit_period = 'year'
            elif _period_meta.get('source') == 'recent_program':
                visit_period = 'recent_program'
            else:
                visit_period = 'month'
        visit_data = _fetch_visit_data(customer_code, dfv, visit_period, _period_meta.get('month_label'), _period_meta.get('year_label')) if (not dfv.empty) else {'current_month': {'total': 0, 'details': {}}, 'current_year': {'total': 0, 'details': {}}}
        visit_debug['month_total'] = int(visit_data['current_month']['total'])
        visit_debug['year_total'] = int(visit_data['current_year']['total'])
    else:
        # لا يوجد سجل زيارات مخزن
        visit_data = {'current_month': {'total': 0, 'details': {}}, 'current_year': {'total': 0, 'details': {}}}
        visit_debug['reason'] = 'empty_visit_history'

    # إلحاق تشخيص الماكينات الأساسية ضمن تشخيص الزيارات لعرضه في الواجهة
    if isinstance(visit_debug, dict):
        visit_debug['primary'] = primary_debug

    # 11. إرجاع الهيكل المطلوب من الواجهة
    return {
        'success': True,
        'message': message,
        'visit_period': visit_period,
        'customer_data': customer_data,
        'dynamic_fields': dynamic_fields,
        'primary_record': primary_record,
        'primary_match_mode': primary_match_mode,
        'branch_section': branch_section,
        'visit_data': visit_data,
        'visit_debug': visit_debug,
        'serial_list': serial_list,
        'cols': list(mapped_df.columns),
        # إضافة نتيجة التجميع بالكامل للسماح للـ JS بالتعامل مع الكيانات المتعددة إذا لزم الأمر
        'items': grouped_nested_results, 
    }

@machine_reports_bp.route('/inquiry', methods=['GET'])
@login_required
@role_required(['admin', 'data_entry', 'user'])
@permission_required('can_inquiry')
def inquiry_view():
    """عرض صفحة الاستعلام"""
    try:
        from models import User
        maintenance_names = [u.username for u in User.query.order_by(User.username).all()]
    except Exception:
        maintenance_names = []
    return render_template('inquiry_popup.html', categories=CATEGORIES, maintenance_names=maintenance_names)


@machine_reports_bp.route('/api/inquiry_search', methods=['POST'])
@login_required
@role_required(['admin', 'data_entry', 'user'])
@permission_required('can_inquiry')
def api_inquiry_search():
    """واجهة API لتنفيذ البحث بناءً على مدخلات المستخدم"""
    data = request.json
    category = data.get('category')
    search_type = data.get('search_type')
    query = data.get('query')
    visit_period = (data.get('visit_period') or 'recent_program').strip()

    if category not in CATEGORIES or search_type not in ['code', 'serial', 'name', 'machine_code'] or not query:
        return jsonify({'success': False, 'message': 'بيانات بحث غير صالحة.', 'items':[]}), 400

    if visit_period not in ['month','year','recent_program']:
        visit_period = 'recent_program'
    result = _inquiry_search(category, search_type, query, visit_period)

    return jsonify(result)


@machine_reports_bp.route('/api/service_tickets/save', methods=['POST'])
@login_required
@role_required(['admin', 'data_entry', 'user'])
def api_save_service_tickets():
    """حفظ تذاكر الخدمات المرتبطة بكل مسلسل حسب تبويب الاستعلام."""
    payload = request.json or {}
    category_key = payload.get('category')
    tickets = payload.get('tickets') or []
    customer_data = payload.get('customer_data') or {}

    if category_key not in CATEGORIES:
        return jsonify({'success': False, 'message': 'قسم غير صالح.', 'errors': ['قسم غير صالح']}), 400

    if not isinstance(tickets, list) or len(tickets) == 0:
        return jsonify({'success': False, 'message': 'لا توجد سجلات للحفظ.', 'errors': ['لا توجد سجلات']}), 400

    errors = []
    now = datetime.utcnow()
    category_label = CATEGORIES.get(category_key, category_key)

    # تحقق رقم الإذن: أرقام فقط ولكن إلزامي فقط إذا تم اختيار نوع/أنواع عطل
    local_orders = []
    for i, t in enumerate(tickets):
        raw_fts = t.get('fault_types')
        fts = []
        if isinstance(raw_fts, list):
            fts = [x.strip() for x in raw_fts if x and str(x).strip()]
        single_ft = (t.get('fault_type') or '').strip()
        has_fault = (len(fts) > 0) or bool(single_ft)

        on = str(t.get('order_number','')).strip()
        if has_fault:
            # إذا كان هناك عطل، يجب أن يكون رقم الإذن موجودًا وأرقام فقط
            if not on or not on.isdigit():
                errors.append(f"سطر {i+1}: رقم الإذن يجب أن يكون أرقام فقط ومطلوب عند تسجيل عطل.")
        # نجمع الأوامر فقط للمدقق الخاص بتكرار الرقم عبر العملاء المختلفين
        if on:
            local_orders.append(on)

    # تحقق من أنواع الأعطال (يدعم متعددة) + السماح بعطل لماكينة واحدة فقط
    fault_rows_count = 0
    for i, t in enumerate(tickets):
        # تنظيف القيم الفارغة من الأعطال المتعددة قبل التحقق
        raw_fts = t.get('fault_types')
        fts = []
        if isinstance(raw_fts, list):
            fts = [x.strip() for x in raw_fts if x and str(x).strip()]

        if isinstance(fts, list) and len(fts) > 0:
            fault_rows_count += 1
            bads = [x for x in fts if x not in ALLOWED_FAULT_TYPES]
            if bads:
                errors.append(f"سطر {i+1}: نوع/أنواع عطل غير مسموح ({', '.join(bads)}).")
        else:
            # في حالة عدم اختيار أعطال متعددة، نعود للتحقق من حقل واحد اختياري
            ft = (t.get('fault_type') or '').strip()
            if ft and ft not in ALLOWED_FAULT_TYPES:
                errors.append(f"سطر {i+1}: نوع عطل غير مسموح ({ft}).")
            if ft:
                fault_rows_count += 1

    # السماح بتسجيل أعطال لعدة ماكينات: إزالة القيد السابق
    # يمكن للمستخدم تسجيل أكثر من عطل في نفس العملية دون تقييد بعدد واحد

    # السماح بتكرار رقم الإذن فقط إذا تطابق "رقم العميل" و"اسم العميل" مع السجل السابق
    customer_code = customer_data.get('رقم العميل') or customer_data.get('رقم المخبز') or ''
    customer_name = customer_data.get('اسم العميل') or customer_data.get('اسم المخبز') or ''
    if not errors and local_orders:
        try:
            existing_rows = (
                ServiceTicket.query
                .filter(ServiceTicket.order_number.in_(local_orders))
                .all()
            )
            for row in existing_rows:
                same_code = (row.customer_code or '') == (customer_code or '')
                same_name = (row.customer_name or '') == (customer_name or '')
                if not (same_code and same_name):
                    errors.append(f"رقم الإذن مستخدم لعميل آخر: {row.order_number}")
        except Exception:
            pass

    if errors:
        return jsonify({'success': False, 'message': 'فشل الحفظ بسبب أخطاء.', 'errors': errors}), 400

    # حفظ السجلات
    try:
        saved_payload_rows = []
        for t in tickets:
            faults_list = t.get('fault_types') or []
            fault_str = ','.join([x.strip() for x in faults_list if x and x.strip()]) if faults_list else (t.get('fault_type') or '').strip()
            # لا نحفظ السطر إذا لم يُسجَّل أي عطل
            if not fault_str:
                continue
            st = ServiceTicket(
                created_at=now,
                category_key=category_key,
                category_label=category_label,
                fault_type=fault_str,
                order_number=str(t.get('order_number') or '').strip(),
                username=getattr(current_user, 'username', 'unknown'),
                customer_code=customer_code,
                customer_name=customer_data.get('اسم العميل') or customer_data.get('اسم المخبز') or '',
                machine_code=t.get('machine_code') or '',
                machine_serial=t.get('machine_serial') or '',
                main_sub=t.get('main_sub') or '',
                status=t.get('status') or '',
                sim1=t.get('sim1') or '',
                sim2=t.get('sim2') or '',
                services=getattr(current_user, 'username', 'unknown'),
                maintenance=str(t.get('maintenance') or '').strip()
            )
            db.session.add(st)
            # نبني صفًا للترحيل إلى خدمات التجار (المترددين - البيانات الحديثة)
            # إنشاء صف البيانات الأساسية
            row_data = {
                'التاريخ': now.strftime('%Y-%m-%d %H:%M:%S'),
                'القسم': category_label,
                'الادارة': customer_data.get('الادارة') or '',
                'المكتب': customer_data.get('المكتب') or '',
                'رقم العميل': customer_code,
                'اسم العميل': customer_data.get('اسم العميل') or customer_data.get('اسم المخبز') or '',
                'رقم الماكينة': t.get('machine_code') or '',
                'مسلسل': t.get('machine_serial') or '',
                'رئيسية/فرعية': t.get('main_sub') or '',
                'حالة الماكينة': t.get('status') or '',
                'شريحة1': t.get('sim1') or '',
                'شريحة2': t.get('sim2') or '',
                'رقم الإذن': str(t.get('order_number') or '').strip(),
                'الحوالة المطلوبة': str(t.get('required_transfer') or '').strip(),
                'القائم بالصيانة': str(t.get('maintenance') or '').strip(),
                'اسم المستخدم': getattr(current_user, 'username', 'unknown'),
                'خدمات': getattr(current_user, 'username', 'unknown'),
                'ملاحظات1': str(t.get('notes1') or '').strip()
            }
            
            # إضافة أعمدة الأعطال المنفصلة مع القيم الافتراضية فارغة
            fault_types_in_ticket = [x.strip() for x in faults_list if x and x.strip()]
            for fault_type in ALLOWED_FAULT_TYPES:
                row_data[fault_type] = '1' if fault_type in fault_types_in_ticket else ''
            
            saved_payload_rows.append(row_data)
        
        # التحقق من صحة البيانات قبل الحفظ النهائي
        db.session.flush()
        db.session.commit()

        # ترحيل البيانات إلى قسم خدمات التجار — المترددين: البيانات الحديثة (البرنامج)
        try:
            if saved_payload_rows:
                import pandas as pd
                from models_reports import ReportState
                # تحويل إلى DataFrame بنفس تنسيق التقارير
                df = pd.DataFrame(saved_payload_rows)

                # تطبيق فلترة السجلات: الاحتفاظ فقط بما يحتوي على أعطال أو تاريخ أو خدمات أو صيانة
                fault_cols = [c for c in ALLOWED_FAULT_TYPES if c in df.columns]
                has_fault = df[fault_cols].eq('1').any(axis=1) if fault_cols else pd.Series([False] * len(df))

                def _nonempty(df, col):
                    return df[col].astype(str).str.strip() != '' if col in df.columns else pd.Series([False] * len(df))

                mask = has_fault | _nonempty(df, 'القائم بالصيانة') | _nonempty(df, 'خدمات') | _nonempty(df, 'التاريخ')
                df_filtered = df[mask]

                # إزالة التكرار اعتمادًا على "رقم الإذن" إن وجد، وإلا على الصف بالكامل
                if not df_filtered.empty:
                    if 'رقم الإذن' in df_filtered.columns:
                        df_filtered = df_filtered.drop_duplicates(subset=['رقم الإذن'], keep='last')
                    else:
                        df_filtered = df_filtered.drop_duplicates(keep='last')

                # حفظ/تحديث سجل الحالة للأدمن الحالي مع الدمج التراكمي لسجلات شاشة الاستعلام فقط
                user_id = getattr(current_user, 'id', None)
                row = ReportState.query.filter_by(category='trader_frequent:recent_program', user_id=user_id).first()
                if not row:
                    row = ReportState(category='trader_frequent:recent_program', user_id=user_id)
                    db.session.add(row)
                    existing_df = pd.DataFrame()
                else:
                    try:
                        existing_df = _json_to_df(row.data_json) if row.data_json else pd.DataFrame()
                    except Exception:
                        existing_df = pd.DataFrame()

                # الملف يبدأ فارغًا، وأول حفظ يضيف أول سجل؛ ثم تتراكم السجلات الجديدة فقط
                combined = pd.concat([existing_df, df_filtered], ignore_index=True) if not existing_df.empty else df_filtered
                if not combined.empty:
                    if 'رقم الإذن' in combined.columns:
                        combined = combined.drop_duplicates(subset=['رقم الإذن'], keep='last')
                    else:
                        combined = combined.drop_duplicates(keep='last')

                row.data_json = _df_to_json(combined) if not combined.empty else _df_to_json(pd.DataFrame())
                db.session.commit()
        except Exception as ex2:
            # لا نفشل الحفظ الرئيسي بسبب مشكلة ترحيل البيانات المساعدة
            current_app = None
            try:
                from flask import current_app as _ca
                current_app = _ca
            except Exception:
                pass
            try:
                if current_app:
                    current_app.logger.warning(f'Trader frequent recent_program sync warning: {ex2}')
            except Exception:
                pass
    except Exception as ex:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'فشل الحفظ: {ex}', 'errors': [str(ex)]}), 500

    return jsonify({'success': True, 'message': 'تم حفظ السجلات بنجاح.', 'saved': len(tickets)})

@machine_reports_bp.route('/api/recent_program/reset', methods=['POST'])
@login_required
@role_required(['admin'])
def api_reset_recent_program():
    """تفريغ جدول recent_program وإنشاؤه بنفس أسماء الأعمدة ولكن فارغ لبدء استقبال البيانات الجديدة."""
    try:
        import pandas as pd
        from models_reports import ReportState

        # الأعمدة الأساسية المستخدمة في شاشة الاستعلام + أعمدة الأعطال
        base_cols = [
            'الادارة','المكتب','رقم العميل','اسم العميل','رقم الماكينة','مسلسل',
            'رئيسية/فرعية','حالة الماكينة','شريحة1','شريحة2','رقم الإذن',
            'الحوالة المطلوبة','القائم بالصيانة','اسم المستخدم','خدمات','ملاحظات1'
        ]
        fault_cols = list(ALLOWED_FAULT_TYPES)
        cols = base_cols + fault_cols

        # إنشاء DataFrame فارغ بهذه الأعمدة
        df_empty = pd.DataFrame(columns=cols)

        # تحديث سجل الحالة للمستخدم الحالي
        user_id = getattr(current_user, 'id', None)
        row = ReportState.query.filter_by(category='trader_frequent:recent_program', user_id=user_id).first()
        if not row:
            row = ReportState(category='trader_frequent:recent_program', user_id=user_id)
            db.session.add(row)

        row.data_json = _df_to_json(df_empty)
        db.session.commit()

        return jsonify({'success': True, 'message': 'تم تفريغ recent_program وإنشاؤه فارغًا بنفس الأعمدة.'})
    except Exception as ex:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'فشل التفريغ: {ex}'}), 500

@machine_reports_bp.route("/")
@login_required
@permission_required('can_general_reports')
def index():
    return render_template("reports/index.html", categories=CATEGORIES, title="التقارير العامة")

@machine_reports_bp.route("/<category>", methods=["GET"])
@login_required
@permission_required('can_general_reports')
def category_view(category):
    if category not in CATEGORIES:
        flash("قسم غير موجود", "warning")
        return redirect(url_for("machine_reports_bp.index"))

    row = _load_state(category)
    is_admin = getattr(current_user, "role", None) == "admin"

    df = _json_to_df(row.data_json) if row and row.data_json else pd.DataFrame()
    mapping = json.loads(row.mapping_json) if row and row.mapping_json else {}

    q = request.args.get("q", "", type=str)
    search_in = request.args.get("search_in", "all")
    page = request.args.get("page", 1, type=int)
    page_size = request.args.get("page_size", 25, type=int)
    page_size = 10 if page_size < 10 else 1000 if page_size > 1000 else page_size

    search_cols_for_view = [search_in] if search_in != 'all' else None

    if not df.empty:
        mapped    = _apply_mapping(df, mapping)
        filtered = _filter_dataframe(mapped, q, search_cols=search_cols_for_view)
        visible  = _drop_empty_columns(filtered)
        
        if visible.empty:
            page_df = pd.DataFrame(); total_pages = 1; search_cols = []
        else:
            page_df, total = _paginate(visible, page, page_size)
            total_pages = max(1, (total + page_size - 1)//page_size)
            search_cols = list(visible.columns)
    else:
        page_df = pd.DataFrame(); total_pages = 1; search_cols = []

    cols = list(page_df.columns)
    rows = page_df.to_dict(orient="records") if not page_df.empty else [] 

    def _page_url(n):
        return url_for('machine_reports_bp.category_view', category=category, q=q, search_in=search_in, page=n, page_size=page_size)
    first_pages = [n for n in [1,2,3] if n <= total_pages]
    pagination = {
        "prev": _page_url(page-1) if page>1 else None,
        "next": _page_url(page+1) if page<total_pages else None,
        "first_pages": [{"n":n,"url":_page_url(n),"active":(n==page)} for n in first_pages],
        "show_ellipsis": total_pages > 3,
        "last": {"n":total_pages,"url":_page_url(total_pages),"active":(page==total_pages)} if total_pages>3 else None,
        "page": page, "total_pages": total_pages
    }

    try:
        # التحقق مما إذا كانت الدالة موجودة للسماح بتمكين/تعطيل زر الحفظ
        _ = url_for('machine_reports_bp.save_mapping', category=category) 
        mapping_enabled = True
    except Exception:
        mapping_enabled = False

    return render_template("reports/category.html",
                           categories=CATEGORIES,
                           category=category,
                           category_label=CATEGORIES[category],
                           has_data=(not df.empty),
                           q=q, search_in=search_in, page=page, page_size=page_size,
                           cols=cols, rows=rows,
                           order_csv=",".join(mapping.get("order", [])),
                           rename_lines="\n".join([f"{k}=>{v}" for k,v in (mapping.get('rename') or {}).items()]),
                           is_admin=is_admin, mapping_enabled=mapping_enabled,
                           pagination=pagination, search_cols=search_cols)

@machine_reports_bp.route("/<category>/import_view", methods=["GET"])
@login_required
@role_required(['admin', 'data_entry'])
@permission_required('can_general_reports')
def import_view(category):
    if category not in CATEGORIES:
        flash("قسم غير موجود", "warning")
        return redirect(url_for("machine_reports_bp.index"))
    return render_template("reports/import.html",
                           categories=CATEGORIES,
                           category=category,
                           category_label=CATEGORIES[category])

@machine_reports_bp.route("/<category>/save_mapping", methods=["POST"])
@login_required
@role_required(['admin']) 
def save_mapping(category):
    if category not in CATEGORIES:
        flash("قسم غير موجود", "warning")
        return redirect(url_for("machine_reports_bp.index"))

    order_csv = (request.form.get("order_csv") or "").strip()
    order = [c.strip() for c in order_csv.split(",") if c.strip()] if order_csv else []

    rename_lines = (request.form.get("rename_lines") or "").strip()
    rename = {}
    if rename_lines:
        for line in rename_lines.splitlines():
            if "=>" in line:
                old, new = line.split("=>", 1)
                old, new = old.strip(), new.strip()
                if old and new:
                    rename[old] = new

    mapping = {"order": order, "rename": rename}
    _save_state(category, df=None, mapping=mapping)
    # تحديث الكاش بعد تغيير المابنج
    try:
        _invalidate_inquiry_cache(category)
    except Exception:
        pass
    flash("تم حفظ إعدادات المابنج.", "success")
    return redirect(url_for("machine_reports_bp.category_view", category=category))

@machine_reports_bp.route("/<category>/import", methods=["POST"])
@login_required
@role_required(['admin', 'data_entry'])
@permission_required('can_general_reports')
def import_files(category):
    if category not in CATEGORIES:
        flash("قسم غير موجود", "warning")
        return redirect(url_for("machine_reports_bp.index"))

    # 1. تجميع الملفات المرفوعة ومعلوماتها
    uploaded_files_info = []
    for i in range(1, MAX_FILES + 1):
        file_storage = request.files.get(f"file{i}")
        if file_storage and file_storage.filename:
            uploaded_files_info.append({"file_storage": file_storage, "filename": file_storage.filename, "index": i})

    if len(uploaded_files_info) == 0:
        flash("برجاء اختيار ملف (المسلسلات) على الأقل.", "warning")
        return redirect(url_for("machine_reports_bp.import_view", category=category))

    # 2. قراءة الملفات وتجميع البيانات التي تم قراءتها بنجاح
    dfs_all = [] 
    successful_files = []
    failed_filenames = []
    
    for item in uploaded_files_info:
        try:
            df = _read_any(item["file_storage"])
            dfs_all.append(df)
            if df.empty:
                failed_filenames.append(f'{item["filename"]} (الموقع: ملف {item["index"]} - فارغ/فشل في القراءة)')
            else:
                successful_files.append(item["filename"])
        except Exception as ex:
             current_app.logger.exception(f"Error reading file {item['filename']}: {ex}")
             dfs_all.append(pd.DataFrame()) 
             failed_filenames.append(f'{item["filename"]} (الموقع: ملف {item["index"]} - فشل حاد في القراءة)')


    # 3. الدمج
    try:
        out_df = _merge_all(dfs_all, category)
        
        if out_df.empty and successful_files:
             flash("تم قراءة الملفات، لكن عملية الدمج لم تنتج عنها سجلات صالحة.", "warning")
             return redirect(url_for("machine_reports_bp.category_view", category=category))
        
        # 💡 تم التعديل: هنا يتم استدعاء دالة الحفظ التي تستخدم user_id
        _save_state(category, df=out_df)
        # إبطال الكاش لضمان إعادة بناء الفهارس مع البيانات الجديدة
        try:
            _invalidate_inquiry_cache(category)
        except Exception:
            pass
        
        msg = f"تم استيراد ودمج {len(successful_files)} ملف(ات) بنجاح. إجمالي السجلات بعد الدمج: {len(out_df)}"
        if failed_filenames:
             msg += f". ملاحظة: لم يتم استخدام/قراءة الملفات التالية: {', '.join(failed_filenames)}"
             flash(msg, "warning")
        else:
             flash(msg, "success")
             
    except Exception as ex:
        current_app.logger.exception("Import error:")
        flash(f"خطأ في دمج الملفات: {ex}", "danger")

    return redirect(url_for("machine_reports_bp.category_view", category=category))

@machine_reports_bp.route("/<category>/export", methods=["GET"])
@login_required
def export_excel(category):
    if category not in CATEGORIES:
        flash("قسم غير موجود", "warning")
        return redirect(url_for("machine_reports_bp.index"))

    # 💡 تم التعديل: هنا يتم استدعاء دالة التحميل التي تستخدم user_id
    row = _load_state(category)
    if not row or not row.data_json:
        flash("لا توجد بيانات لتصديرها. برجاء الاستيراد أولاً بواسطة الأدمن.", "warning")
        return redirect(url_for("machine_reports_bp.category_view", category=category))

    df = _json_to_df(row.data_json)
    mapping = json.loads(row.mapping_json) if row and row.mapping_json else {}
    q = request.args.get("q", "", type=str)
    search_in = request.args.get("search_in", "all")

    out = _apply_mapping(df, mapping)
    search_cols_for_export = [search_in] if search_in != 'all' else None
    out = _filter_dataframe(out, q, search_cols=search_cols_for_export)
    out = _drop_empty_columns(out)
    out = _coerce_text_df(out)

    output = io.BytesIO()
    try:
        import xlsxwriter
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            sheet = "data"
            out.to_excel(writer, index=False, sheet_name=sheet)
            ws = writer.sheets[sheet]
            book = writer.book
            header_fmt = book.add_format({"bold": True, "bg_color": "#E2E8F0", "align": "center", "valign": "vcenter", "border": 1, "num_format": "@"})
            cell_fmt   = book.add_format({"align": "center", "valign": "vcenter", "border": 1, "num_format": "@"})
            for i, c in enumerate(out.columns):
                ws.write(0, i, c, header_fmt)
            n_rows, n_cols = len(out.index), len(out.columns)
            if n_rows > 0 and n_cols > 0:
                from xlsxwriter.utility import xl_rowcol_to_cell
                start = xl_rowcol_to_cell(1, 0); end = xl_rowcol_to_cell(n_rows, n_cols-1)
                ws.conditional_format(f"{start}:{end}", {"type":"no_blanks", "format": cell_fmt})
            for i, c in enumerate(out.columns):
                series = out[c].astype(str)
                width = min(max([len(str(c))] + [len(s) for s in series.tolist()]) + 2, 60)
                ws.set_column(i, i, width)
            ws.freeze_panes(1, 0)
        output.seek(0)
    except ModuleNotFoundError:
        try:
            from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
            from openpyxl.utils import get_column_letter
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                sheet = "data"
                out.to_excel(writer, index=False, sheet_name=sheet)
                ws = writer.sheets[sheet]
                header_fill = PatternFill("solid", fgColor="E2E8F0")
                header_font = Font(bold=True)
                thin = Side(border_style="thin", color="CCCCCC")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                center = Alignment(horizontal="center", vertical="center", wrap_text=False)
                for cell in ws[1]:
                    cell.fill = header_fill; cell.font = header_font
                    cell.alignment = center; cell.border = border; cell.number_format = "@"
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        if cell.value not in (None, ""):
                            cell.alignment = center; cell.border = border; cell.number_format = "@"
                for i, c in enumerate(out.columns, start=1):
                    series = out[c].astype(str).tolist() if not out.empty else []
                    width = min(max([len(str(c))] + [len(s) for s in series]) + 2, 60) if series else len(str(c))
                    ws.column_dimensions[get_column_letter(i)].width = width
                ws.freeze_panes = "A2"
            output.seek(0)
        except Exception:
             # في حال فشل الاستيراد والتنسيق، نكتفي بكتابة ملف Excel بدون تنسيق
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                out.to_excel(writer, index=False, sheet_name="data")
            output.seek(0)


    filename = f"{CATEGORIES[category]}_export.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def _paginate(df: pd.DataFrame, page: int, page_size: int):
    """وظيفة التقسيم للصفحات"""
    n = len(df)
    start = max(0, (page - 1) * page_size)
    end = start + page_size
    return df.iloc[start:end], n

# ==================== (4) وظائف الدمج (Merge) ====================

_AR_DIGITS = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")
def _norm_key_text(s: str) -> str:
    if s is None:
        return ""
    t = str(s).translate(_AR_DIGITS)
    # توحيد أشكال الألف، إزالة التطويل
    t = t.replace("\u0640", "").replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
    # تحسين التطبيع للأحرف الشائعة في الأسماء العربية
    # تحويل الياء المقصورة إلى ياء عادية، والتاء المربوطة إلى هاء لتقليل الفروق
    t = t.replace("ى", "ي").replace("ة", "ه").replace("ئ", "ي").replace("ؤ", "و")
    # إزالة المسافات الزائدة
    t = re.sub(r"\s+", " ", t).strip()
    return t

def _normalize_key_cols(df: pd.DataFrame, join_cols: list[str]) -> pd.DataFrame:
    out = df.copy()
    for c in join_cols:
        if c in out.columns:
            out[c] = out[c].map(lambda v: _norm_key_text("" if (pd.isna(v) or v is None) else str(v)))
        else:
            out[c] = ""
    return out

def _pick_entity_keys(category: str, dfs):
    """يحدد المفاتيح الأكثر احتمالاً للدمج بناءً على المفاتيح الموجودة في الملفات المرفوعة."""
    
    # قائمة تفضيلات مفاتيح الدمج بترتيب الأولوية
    prefs_order = [("رقم العميل", "اسم العميل")]
    
    # إضافة مفاتيح الفئة والافتراضية القديمة كبدائل، مع تجنب التكرار
    prefs_order.extend(ENTITY_KEYS.get(category, []))
    prefs_order.extend(ENTITY_KEYS["default"])
    
    # إزالة التكرارات مع الحفاظ على الترتيب
    seen = set()
    unique_prefs = []
    for pair in prefs_order:
        if pair not in seen:
            seen.add(pair)
            unique_prefs.append(pair)
    
    available = set()
    for d in dfs:
        available.update(list(d.columns)) 
        
    for a, b in unique_prefs:
        # نبحث عن تطابق على الأقل في أحد المفتاحين (الرقم أو الاسم)
        keys = [k for k in [a, b] if k in available]
        if keys:
            return keys
            
    return []

def _pick_office_keys(dfs):
    """تحديد مفاتيح الدمج لمستوى المكتب/الإدارة مع دعم المرادفات العربية والإنجليزية.
    يعيد قائمة بالمفاتيح الموجودة فعليًا (1 أو 2) بحسب الأعمدة المتاحة.
    """
    office_pairs = [
        ("الادارة", "المكتب"),
        ("الإدارة", "المكتب"),
        ("المركز", "المكتب"),
        ("المحافظة", "المكتب"),
        ("Administration", "Office"),
        ("Admin", "Office"),
        ("Department", "Office"),
        ("Branch", "Office"),
    ]

    # إزالة التكرارات مع الحفاظ على الترتيب
    seen = set(); unique_pairs = []
    for pair in office_pairs:
        if pair not in seen:
            seen.add(pair); unique_pairs.append(pair)

    available = set()
    for d in dfs:
        available.update(list(d.columns))

    # اختر أول زوج تتوفر منه مفاتيح فعلية في البيانات
    for a, b in unique_pairs:
        keys = [k for k in [a, b] if k in available]
        if keys:
            return keys
    # fallback: أي عمود يحتوي كلمة مكتب/ادارة
    office_like = [c for c in available if any(tok in str(c).lower() for tok in ["office","branch","اداره","إداره","الاداره","الإدارة","مكتب"])]
    if office_like:
        return [office_like[0]]
    return []

def _read_any(file_storage) -> pd.DataFrame:
    if not file_storage or not file_storage.filename: return pd.DataFrame()
    name = file_storage.filename.lower()
    bio = io.BytesIO(file_storage.read())
    df = pd.DataFrame()
    try:
        if name.endswith((".xlsx",".xls")):
            df = pd.read_excel(bio, dtype=str)
        else:
            df = pd.read_csv(bio, dtype=str, encoding="utf-8", errors="ignore")
    except Exception:
        try:
            bio.seek(0); df = pd.read_excel(bio, dtype=str)
        except Exception:
            bio.seek(0); df = pd.read_csv(bio, dtype=str, encoding="utf-8", errors="ignore")
            
    return _coerce_text_df(df)


def _left_enrich(base: pd.DataFrame, data: pd.DataFrame, keys: list[str], suffix="__D") -> pd.DataFrame:
    if data is None or data.empty: 
        return base.copy()
    if not keys:
        return base.copy()

    # 1. التنظيف وتوحيد المفاتيح
    baseN  = _normalize_key_cols(base, keys)
    dataN  = _normalize_key_cols(data, keys)

    # 2. إزالة التكرارات من بيانات الإثراء
    dataN = dataN.drop_duplicates(subset=keys, keep="first")

    # 3. قائمة الأعمدة غير المفتاحية في بيانات الإثراء
    enrich_cols = [c for c in dataN.columns if c not in keys]

    # 4. دمج الـ DataFrame (سيعطي أسماء أعمدة بلاحقة __D للأعمدة المكررة غير المفتاحية)
    merged = baseN.merge(dataN, how="left", on=keys, suffixes=("", suffix), copy=True)
    
    # 5. دمج البيانات من الأعمدة ذات اللاحقة (Enrichment) إلى الأعمدة الأساسية (Base) إذا كانت فارغة
    for col in enrich_cols:
        col_with_suffix = f"{col}{suffix}"
        if col_with_suffix in merged.columns and col in merged.columns:
            lvals = merged[col].astype(str).map(_textify)
            rvals = merged[col_with_suffix].astype(str).map(_textify)
            
            merged[col] = lvals.where(lvals != "", rvals)
            
            merged.drop(columns=[col_with_suffix], inplace=True)
        elif col_with_suffix in merged.columns and col not in merged.columns:
             merged.rename(columns={col_with_suffix: col}, inplace=True)

    # 6. إزالة أي أعمدة مكررة قد تكون بقيت (مثلاً إذا كان اسم العمود في BaseN مكررًا أصلاً)
    return _coerce_text_df(merged)

def _append_unmatched(base: pd.DataFrame, data: pd.DataFrame, keys: list[str]) -> pd.DataFrame:
    if data is None or data.empty or not keys:
        return base.copy()

    baseK = _normalize_key_cols(base, keys)
    dataK = _normalize_key_cols(data, keys)

    base_keys = set(tuple(row) for row in baseK[keys].itertuples(index=False, name=None))
    data_only = dataK[~dataK[keys].apply(lambda r: tuple(r) in base_keys, axis=1)]

    if data_only.empty:
        return base.copy()

    union_cols = list(dict.fromkeys(list(base.columns) + [c for c in data_only.columns if c not in base.columns]))
    out = pd.concat([
        base.reindex(columns=union_cols),
        data_only.reindex(columns=union_cols)
    ], ignore_index=True)

    return _coerce_text_df(out)
    
def _apply_standard_transformations(df: pd.DataFrame) -> pd.DataFrame:
    """
    تطبيق التحويلات القياسية المطلوبة على الأعمدة مباشرة بعد الاستيراد والدمج.
    مثل تحويل قيم (ماكينة رئيسية/فرعية) من رقمية إلى نصية.
    """
    if df.empty:
        return df

    out = df.copy()
    target_col = 'ماكينة رئيسية/فرعية' 

    if target_col in out.columns:
        
        # 💡 دالة التحويل: تحويل القيمة الرقمية 0/1 إلى نص
        def _convert_primary_secondary(value):
            """تحويل 0 إلى رئيسية و 1 إلى فرعية."""
            # نضمن تحويل القيمة إلى نص نظيف (مثل '0' أو '1') حتى لو كانت أصلها رقم عشري (0.0، 1.0)
            value_str = _textify(value)
            
            if value_str == '0':
                return 'رئيسية'
            elif value_str == '1':
                return 'فرعية'
            # نترك القيمة الأصلية أو قيمة فارغة إذا لم تكن 0 أو 1
            return value_str 
            
        out[target_col] = out[target_col].map(_convert_primary_secondary)
        
    return out

def _merge_all(files: list[pd.DataFrame], category: str) -> pd.DataFrame:
    # لا تقم بتصفية القوائم للحفاظ على ترتيب الملفات (1..6)
    if not files or files[0] is None or files[0].empty:
        return pd.DataFrame()

    base = files[0].copy()
    
    def _has_keys(df: pd.DataFrame, keys: list[str]) -> bool:
        return keys and all(k in df.columns for k in keys)

    # الملفات 2، 3، 4 — دمج بمفاتيح كيان مرنة (رقم/اسم العميل أو المرادفات) حسب الفئة
    for idx in [1, 2, 3]:
        if idx < len(files):
            df = files[idx]
            if df is not None and not df.empty:
                keys = _pick_entity_keys(category, [base, df])
                if _has_keys(base, keys) and _has_keys(df, keys):
                    base = _left_enrich(base, df, keys, suffix=f"__D{idx+1}")

    # الملفات 5، 6 — دمج بمفاتيح الإدارة/المكتب بمرادفات مرنة
    for idx in [4, 5]:
        if idx < len(files):
            df = files[idx]
            if df is not None and not df.empty:
                keys = _pick_office_keys([base, df])
                if _has_keys(base, keys) and _has_keys(df, keys):
                    base = _left_enrich(base, df, keys, suffix=f"__D{idx+1}")

    # تطبيق التحويلات القياسية بعد الدمج
    base = _apply_standard_transformations(base)
    return _coerce_text_df(base)
# ==================== (1.a) تسريع الاستعلام: كاش وفهارس في الذاكرة ====================
# كاش الاستعلام لكل تبويب (category): يحتفظ بنسخة DataFrame بعد المابنج + فهارس سريعة
INQUIRY_CACHE: dict[str, dict] = {}

def _mapping_signature(mapping: dict) -> str:
    try:
        import hashlib, json as _json
        s = _json.dumps(mapping or {}, ensure_ascii=False, sort_keys=True)
        return hashlib.sha1(s.encode('utf-8')).hexdigest()
    except Exception:
        return str(len(mapping or {}))

def _build_inquiry_cache(category: str):
    """إنشاء/تحديث كاش الاستعلام لفئة معينة: DataFrame بعد المابنج + فهارس للبحث."""
    row = _load_state(category)
    df = _json_to_df(row.data_json) if (row and row.data_json) else pd.DataFrame()
    mapping = json.loads(row.mapping_json) if (row and row.mapping_json) else {}
    mapped_df = _apply_mapping(df, mapping)
    mapped_df = _drop_empty_columns(mapped_df)

    indexes = {
        'code': {},          # رقم العميل/المخبز/التاجر/الماكينة (مطابقة كاملة)
        'serial': {},        # المسلسل ومترادفاته (مطابقة كاملة)
        'name': {},          # الاسم الكامل (مطابقة كاملة)
        'machine_code': {},  # رقم/كود الماكينة (مطابقة كاملة)
        # مطابقة جزئية (بادئة) لأكواد ومسلسلات وكود الماكينة
        'code_prefix3': {}, 'code_prefix5': {},
        # مطابقة جزئية (نهاية/لاحقة) للأكواد لتسريع البحث بالأرقام الأخيرة
        'code_suffix3': {}, 'code_suffix5': {},
        'serial_prefix3': {}, 'serial_prefix5': {},
        'machine_code_prefix3': {}, 'machine_code_prefix5': {},
        # فهرسة كلمات الاسم لدعم المطابقة الجزئية على مستوى الكلمات
        'name_token': {}, 'name_token_prefix3': {}, 'name_token_prefix5': {},
    }

    all_cols = list(mapped_df.columns)

    # أعمدة الكود (أولوية + بدائل عامة)
    code_cols = [c for c in ['رقم العميل', 'رقم المخبز', 'رقم التاجر', 'رقم الماكينة'] if c in all_cols]
    if not code_cols:
        code_cols = [c for c in all_cols if ('رقم' in str(c) and 'مسلسل' not in str(c) and 'قومي' not in str(c))]

    # أعمدة المسلسل
    serial_candidates = ['مسلسل الماكينة', 'مسلسل', 'serial', 'pos serial', 'sn', 'رقم الماكينة', 'machine serial']
    serial_cols = [c for c in all_cols if any(tok in str(c).lower() for tok in serial_candidates)]
    if not serial_cols:
        serial_cols = [c for c in all_cols if 'مسلسل' in str(c)]

    # أعمدة رقم/كود الماكينة
    mc_tokens = ['رقم الماكينة', 'كود الماكينة', 'machine code', 'machine id', 'pos id', 'terminal id', 'رقم الجهاز', 'كود الجهاز']
    mc_tokens = [t.lower() for t in mc_tokens]
    machine_code_cols = [c for c in all_cols if any(tok in str(c).lower() for tok in mc_tokens)]
    if not machine_code_cols:
        machine_code_cols = [c for c in all_cols if ('رقم' in str(c) and 'ماك' in str(c))]
    if not machine_code_cols:
        machine_code_cols = [c for c in all_cols if ('code' in str(c).lower() and 'machine' in str(c).lower())]

    # أعمدة الاسم
    name_cols = [c for c in ['اسم العميل', 'اسم المخبز', 'اسم التاجر'] if c in all_cols]
    if not name_cols:
        name_cols = [c for c in all_cols if 'اسم' in str(c)]

    # بناء الفهارس: قيمة مُطبّعة → قائمة فهارس صفوف
    for idx, r in mapped_df.iterrows():
        try:
            # code
            for c in code_cols:
                v = _textify(r.get(c))
                if v:
                    k = _norm_key_text(v)
                    if k:
                        indexes['code'].setdefault(k, []).append(idx)
                        # بادئة 3 و5
                        if len(k) >= 3:
                            p3 = k[:3]; indexes['code_prefix3'].setdefault(p3, []).append(idx)
                            s3 = k[-3:]; indexes['code_suffix3'].setdefault(s3, []).append(idx)
                        if len(k) >= 5:
                            p5 = k[:5]; indexes['code_prefix5'].setdefault(p5, []).append(idx)
                            s5 = k[-5:]; indexes['code_suffix5'].setdefault(s5, []).append(idx)
            # serial
            for c in serial_cols:
                v = _textify(r.get(c))
                if v:
                    k = _norm_key_text(v)
                    if k:
                        indexes['serial'].setdefault(k, []).append(idx)
                        if len(k) >= 3:
                            p3 = k[:3]; indexes['serial_prefix3'].setdefault(p3, []).append(idx)
                        if len(k) >= 5:
                            p5 = k[:5]; indexes['serial_prefix5'].setdefault(p5, []).append(idx)
            # machine_code
            for c in machine_code_cols:
                v = _textify(r.get(c))
                if v:
                    k = _norm_key_text(v)
                    if k:
                        indexes['machine_code'].setdefault(k, []).append(idx)
                        if len(k) >= 3:
                            p3 = k[:3]; indexes['machine_code_prefix3'].setdefault(p3, []).append(idx)
                        if len(k) >= 5:
                            p5 = k[:5]; indexes['machine_code_prefix5'].setdefault(p5, []).append(idx)
            # name
            for c in name_cols:
                v = _textify(r.get(c))
                if v:
                    k = _norm_key_text(v)
                    if k:
                        indexes['name'].setdefault(k, []).append(idx)
                        # فهرسة كلمات الاسم
                        tokens = [t for t in k.split(' ') if t]
                        for t in tokens:
                            indexes['name_token'].setdefault(t, []).append(idx)
                            if len(t) >= 3:
                                pt3 = t[:3]; indexes['name_token_prefix3'].setdefault(pt3, []).append(idx)
                            if len(t) >= 5:
                                pt5 = t[:5]; indexes['name_token_prefix5'].setdefault(pt5, []).append(idx)
        except Exception:
            # نتجاهل أي صف يسبب خطأ في التطبيع/الفهرسة
            pass

    INQUIRY_CACHE[category] = {
        'df': mapped_df,
        'indexes': indexes,
        'state_id': getattr(row, 'id', None),
        'updated_at': getattr(row, 'updated_at', None),
        'mapping_signature': _mapping_signature(mapping),
        'cols': all_cols,
    }

def _get_inquiry_cache(category: str) -> dict:
    """إرجاع كاش صالح؛ يعيد البناء إذا كان غير موجود أو قديم."""
    row = _load_state(category)
    if not row or not row.data_json:
        return {'df': pd.DataFrame(), 'indexes': {'code':{},'serial':{},'name':{},'machine_code':{}}, 'cols': []}
    mapping = json.loads(row.mapping_json) if (row and row.mapping_json) else {}
    sig = _mapping_signature(mapping)
    cached = INQUIRY_CACHE.get(category)
    if (not cached) or (cached.get('state_id') != getattr(row, 'id', None)) or (cached.get('updated_at') != getattr(row, 'updated_at', None)) or (cached.get('mapping_signature') != sig):
        _build_inquiry_cache(category)
        cached = INQUIRY_CACHE.get(category)
    return cached or {
        'df': pd.DataFrame(),
        'indexes': {
            'code':{},'serial':{},'name':{},'machine_code':{},
            'code_prefix3':{},'code_prefix5':{},
            'code_suffix3':{},'code_suffix5':{},
            'serial_prefix3':{},'serial_prefix5':{},
            'machine_code_prefix3':{},'machine_code_prefix5':{},
            'name_token':{},'name_token_prefix3':{},'name_token_prefix5':{},
        },
        'cols': []
    }

def _invalidate_inquiry_cache(category: str):
    try:
        INQUIRY_CACHE.pop(category, None)
    except Exception:
        pass